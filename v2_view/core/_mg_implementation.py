from datetime import date, datetime
from flask import Blueprint, Response, render_template, request, session, flash, redirect, jsonify, send_file
from flask_session import Session
from xlrd import sheet
from modules.Connections import mysql,sqlite
import Configurations as c
import os
import json
# from views.dcfv2.dashboard.display_dataform import displayform
# from views.dcfv2.spreadsheet import dcf_import_excel as importcsv_mgimplementation
import xlrd
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import pandas as pd


db = mysql(*c.DB_CRED)
db.err_page = 0
def is_on_session(): return ('USER_DATA' in session)


# from flask_babel import Babel, format_currency

app = Blueprint("_mg_implementation",__name__,template_folder='templates')
rapid_sql = mysql(*c.DB_CRED)
api_key = "dtirapid@2025!"

class _main:
    def __init__(self, arg):
        
        super(_main, self).__init__()
        self.arg = arg
    
    @app.route('/mgimplementation/delete', methods=['POST'])
    def mgimplementation_delete():
        id = request.form.get('id')
        type = request.form.get('type')
        if ( type == 'fo'):
            sql = "UPDATE mg_implementation_fo SET isDeleted=1, deletedDate=NOW() WHERE MGI_ID={}".format(id)
        else:
            sql = "UPDATE mg_implementation_shf SET isDeleted=1, deletedDate=NOW() WHERE MGI_ID={}".format(id)
        
        result = rapid_sql.do(sql)
        
        return jsonify(result)
    
        # if (result==0):
        #     return jsonify({'success': False, 'message': 'Unable to remove data.','id': id, 'type': type, 'sql' : sql})
        # else:
        #     return jsonify({'success': True, 'message': 'Successfully removed data.'})

    
    @app.route('/mgimplementation_fo_form/submit', methods=['POST'])
    def submit_fo_form():
        if not is_on_session():
            return redirect("/login")

        if request.form.get("action") == "addnew":
            UploadedBy = session["USER_DATA"][0]['id']
            ImpUnits_RCU = request.form.get("ImpUnits_RCU")
            ImpUnits_PCU = request.form.get("ImpUnits_PCU")
            DIP_Name = request.form.get("DIP_Name")
            DIP_Commodity = request.form.get("DIP_Commodity")
            FOInfo_FOName = request.form.get("FOInfo_FOName")
            FOInfo_Address_Region = request.form.get("FOInfo_Address_Region")
            FOInfo_Address_Province = request.form.get("FOInfo_Address_Province")
            FOInfo_Address_Municipality = request.form.get("FOInfo_Address_Municipality")
            FOInfo_Address_Brgy = request.form.get("FOInfo_Address_Brgy")
            FOInfo_Address_Street = request.form.get("FOInfo_Address_Street")
            FOInfo_Head_Fname = request.form.get("FOInfo_Head_Fname")
            FOInfo_Head_Mname = request.form.get("FOInfo_Head_Mname")
            FOInfo_Head_Lname = request.form.get("FOInfo_Head_Lname")
            FOInfo_Head_Extname = request.form.get("FOInfo_Head_Extname")
            FOInfo_Head_Sex = request.form.get("FOInfo_Head_Sex")
            FOInfo_Head_Sector = request.form.get("FOInfo_Head_Sector")
            AcknowOfFarmExp_TargetArea = request.form.get("AcknowOfFarmExp_TargetArea")
            AcknowOfFarmExp_ItemName = request.form.get("AcknowOfFarmExp_ItemName")
            AcknowOfFarmExp_QtyReceived = request.form.get("AcknowOfFarmExp_QtyReceived")
            AcknowOfFarmExp_DateReceived = request.form.get("AcknowOfFarmExp_DateReceived")
            AcknowOfFarmExp_FarmerRep_Fname = request.form.get("AcknowOfFarmExp_FarmerRep_Fname")
            AcknowOfFarmExp_FarmerRep_Mname = request.form.get("AcknowOfFarmExp_FarmerRep_Mname")
            AcknowOfFarmExp_FarmerRep_Lname = request.form.get("AcknowOfFarmExp_FarmerRep_Lname")
            AcknowOfFarmExp_FarmerRep_Extname = request.form.get("AcknowOfFarmExp_FarmerRep_Extname")
            AcknowOfFarmExp_FarmerRep_Designation = request.form.get("AcknowOfFarmExp_FarmerRep_Designation")
            AcknowOfFarmExp_Witness1_ValidatorName = request.form.get("AcknowOfFarmExp_Witness1_ValidatorName")
            AcknowOfFarmExp_Witness1_ValidationDate = request.form.get("AcknowOfFarmExp_Witness1_ValidationDate")
            AcknowOfFarmExp_Witness1_Designation = request.form.get("AcknowOfFarmExp_Witness1_Designation")
            AcknowOfFarmExp_Witness1_DateWitnessed = request.form.get("AcknowOfFarmExp_Witness1_DateWitnessed")
            AcknowOfFarmExp_Witness2_ValidatorName = request.form.get("AcknowOfFarmExp_Witness2_ValidatorName")
            AcknowOfFarmExp_Witness2_ValidationDate = request.form.get("AcknowOfFarmExp_Witness2_ValidationDate")
            AcknowOfFarmExp_Witness2_Remarks = request.form.get("AcknowOfFarmExp_Witness2_Remarks")
            AcknowOfFarmExp_Witness1Raw_ValidatorName = request.form.get("AcknowOfFarmExp_Witness1Raw_ValidatorName")
            AcknowOfFarmExp_Witness1Raw_ValidationDate = request.form.get("AcknowOfFarmExp_Witness1Raw_ValidationDate")
            AcknowOfFarmExp_Witness1Raw_Designation = request.form.get("AcknowOfFarmExp_Witness1Raw_Designation")
            AcknowOfFarmExp_Witness1Raw_DateWitnessed = request.form.get("AcknowOfFarmExp_Witness1Raw_DateWitnessed")
            AcknowOfFarmExp_Witness2Raw_ValidatorName = request.form.get("AcknowOfFarmExp_Witness2Raw_ValidatorName")
            AcknowOfFarmExp_Witness2Raw_ValidationDate = request.form.get("AcknowOfFarmExp_Witness2Raw_ValidationDate")
            AcknowOfFarmExp_Witness2Raw_Remarks = request.form.get("AcknowOfFarmExp_Witness2Raw_Remarks")
            AcknowOfFarmInten_TargetArea = request.form.get("AcknowOfFarmInten_TargetArea")
            AcknowOfFarmInten_ItemName = request.form.get("AcknowOfFarmInten_ItemName")
            AcknowOfFarmInten_QtyReceived = request.form.get("AcknowOfFarmInten_QtyReceived")
            AcknowOfFarmInten_DateReceived = request.form.get("AcknowOfFarmInten_DateReceived")
            AcknowOfFarmInten_FarmerRep_Fname = request.form.get("AcknowOfFarmInten_FarmerRep_Fname")
            AcknowOfFarmInten_FarmerRep_Mname = request.form.get("AcknowOfFarmInten_FarmerRep_Mname")
            AcknowOfFarmInten_FarmerRep_Lname = request.form.get("AcknowOfFarmInten_FarmerRep_Lname")
            AcknowOfFarmInten_FarmerRep_Extname = request.form.get("AcknowOfFarmInten_FarmerRep_Extname")
            AcknowOfFarmInten_FarmerRep_Designation = request.form.get("AcknowOfFarmInten_FarmerRep_Designation")
            AcknowOfFarmInten_Withness1_ValidatorName = request.form.get("AcknowOfFarmInten_Withness1_ValidatorName")
            AcknowOfFarmInten_Withness1_ValidationDate = request.form.get("AcknowOfFarmInten_Withness1_ValidationDate")
            AcknowOfFarmInten_Withness1_Designation = request.form.get("AcknowOfFarmInten_Withness1_Designation")
            AcknowOfFarmInten_Withness1_DateWitnessed = request.form.get("AcknowOfFarmInten_Withness1_DateWitnessed")
            AcknowOfFarmInten_Withness2_ValidatorName = request.form.get("AcknowOfFarmInten_Withness2_ValidatorName")
            AcknowOfFarmInten_Withness2_ValidationDate = request.form.get("AcknowOfFarmInten_Withness2_ValidationDate")
            AcknowOfFarmInten_Withness2_Remarks = request.form.get("AcknowOfFarmInten_Withness2_Remarks")
            AcknowOfFarmRehab_TargetArea = request.form.get("AcknowOfFarmRehab_TargetArea")
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw = request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw")
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw = request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw")
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer = request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer")
            AcknowOfFarmRehab_DateReceived = request.form.get("AcknowOfFarmRehab_DateReceived")
            AcknowOfFarmRehab_FarmerRep_Fname = request.form.get("AcknowOfFarmRehab_FarmerRep_Fname")
            AcknowOfFarmRehab_FarmerRep_Mname = request.form.get("AcknowOfFarmRehab_FarmerRep_Mname")
            AcknowOfFarmRehab_FarmerRep_Lname = request.form.get("AcknowOfFarmRehab_FarmerRep_Lname")
            AcknowOfFarmRehab_FarmerRep_Extname = request.form.get("AcknowOfFarmRehab_FarmerRep_Extname")
            AcknowOfFarmRehab_FarmerRep_Designation = request.form.get("AcknowOfFarmRehab_FarmerRep_Designation")
            AcknowOfFarmRehab_ValidatorName = request.form.get("AcknowOfFarmRehab_ValidatorName")
            AcknowOfFarmRehab_ValidationDate = request.form.get("AcknowOfFarmRehab_ValidationDate")
            AcknowOfFarmRehab_Remarks = request.form.get("AcknowOfFarmRehab_Remarks")
            AcknowOfProd_TypeOfProdInv = request.form.get("AcknowOfProd_TypeOfProdInv")
            AcknowOfProd_QtyItemReceived = request.form.get("AcknowOfProd_QtyItemReceived")
            AcknowOfProd_DateReceived = request.form.get("AcknowOfProd_DateReceived")
            AcknowOfProd_FarmerRep_Fname = request.form.get("AcknowOfProd_FarmerRep_Fname")
            AcknowOfProd_FarmerRep_Mname = request.form.get("AcknowOfProd_FarmerRep_Mname")
            AcknowOfProd_FarmerRep_Lname = request.form.get("AcknowOfProd_FarmerRep_Lname")
            AcknowOfProd_FarmerRep_Extname = request.form.get("AcknowOfProd_FarmerRep_Extname")
            AcknowOfProd_FarmerRep_Designation = request.form.get("AcknowOfProd_FarmerRep_Designation")
            AcknowOfProd_ValidatorName = request.form.get("AcknowOfProd_ValidatorName")
            AcknowOfProd_ValidationDate = request.form.get("AcknowOfProd_ValidationDate")
            AcknowOfProd_Remarks = request.form.get("AcknowOfProd_Remarks")
            filename = ''

            newMGIFO = ("INSERT INTO mg_implementation_fo (UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,FOInfo_FOName,FOInfo_Address_Region,FOInfo_Address_Province,FOInfo_Address_Municipality,FOInfo_Address_Brgy,FOInfo_Address_Street,FOInfo_Head_Fname,FOInfo_Head_Mname,FOInfo_Head_Lname,FOInfo_Head_Extname,FOInfo_Head_Sex,FOInfo_Head_Sector,AcknowOfFarmExp_TargetArea,AcknowOfFarmExp_ItemName,AcknowOfFarmExp_QtyReceived,AcknowOfFarmExp_DateReceived,AcknowOfFarmExp_FarmerRep_Fname,AcknowOfFarmExp_FarmerRep_Mname,AcknowOfFarmExp_FarmerRep_Lname,AcknowOfFarmExp_FarmerRep_Extname,AcknowOfFarmExp_FarmerRep_Designation,AcknowOfFarmExp_Witness1_ValidatorName,AcknowOfFarmExp_Witness1_ValidationDate,AcknowOfFarmExp_Witness1_Designation,AcknowOfFarmExp_Witness1_DateWitnessed,AcknowOfFarmExp_Witness2_ValidatorName,AcknowOfFarmExp_Witness2_ValidationDate,AcknowOfFarmExp_Witness2_Remarks,AcknowOfFarmExp_Witness1Raw_ValidatorName,AcknowOfFarmExp_Witness1Raw_ValidationDate,AcknowOfFarmExp_Witness1Raw_Designation,AcknowOfFarmExp_Witness1Raw_DateWitnessed,AcknowOfFarmExp_Witness2Raw_ValidatorName,AcknowOfFarmExp_Witness2Raw_ValidationDate,AcknowOfFarmExp_Witness2Raw_Remarks,AcknowOfFarmInten_TargetArea,AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived,AcknowOfFarmInten_DateReceived,AcknowOfFarmInten_FarmerRep_Fname,AcknowOfFarmInten_FarmerRep_Mname,AcknowOfFarmInten_FarmerRep_Lname,AcknowOfFarmInten_FarmerRep_Extname,AcknowOfFarmInten_FarmerRep_Designation,AcknowOfFarmInten_Withness1_ValidatorName,AcknowOfFarmInten_Withness1_ValidationDate,AcknowOfFarmInten_Withness1_Designation,AcknowOfFarmInten_Withness1_DateWitnessed,AcknowOfFarmInten_Withness2_ValidatorName,AcknowOfFarmInten_Withness2_ValidationDate,AcknowOfFarmInten_Withness2_Remarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_FarmerRep_Fname,AcknowOfFarmRehab_FarmerRep_Mname,AcknowOfFarmRehab_FarmerRep_Lname,AcknowOfFarmRehab_FarmerRep_Extname,AcknowOfFarmRehab_FarmerRep_Designation,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_ValidationDate,AcknowOfFarmRehab_Remarks,AcknowOfProd_TypeOfProdInv,AcknowOfProd_QtyItemReceived,AcknowOfProd_DateReceived,AcknowOfProd_FarmerRep_Fname,AcknowOfProd_FarmerRep_Mname,AcknowOfProd_FarmerRep_Lname,AcknowOfProd_FarmerRep_Extname,AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,filename) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".
            format(UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,FOInfo_FOName,FOInfo_Address_Region,FOInfo_Address_Province,FOInfo_Address_Municipality,FOInfo_Address_Brgy,FOInfo_Address_Street,FOInfo_Head_Fname,FOInfo_Head_Mname,FOInfo_Head_Lname,FOInfo_Head_Extname,FOInfo_Head_Sex,FOInfo_Head_Sector,AcknowOfFarmExp_TargetArea,AcknowOfFarmExp_ItemName,AcknowOfFarmExp_QtyReceived,AcknowOfFarmExp_DateReceived,AcknowOfFarmExp_FarmerRep_Fname,AcknowOfFarmExp_FarmerRep_Mname,AcknowOfFarmExp_FarmerRep_Lname,AcknowOfFarmExp_FarmerRep_Extname,AcknowOfFarmExp_FarmerRep_Designation,AcknowOfFarmExp_Witness1_ValidatorName,AcknowOfFarmExp_Witness1_ValidationDate,AcknowOfFarmExp_Witness1_Designation,AcknowOfFarmExp_Witness1_DateWitnessed,AcknowOfFarmExp_Witness2_ValidatorName,AcknowOfFarmExp_Witness2_ValidationDate,AcknowOfFarmExp_Witness2_Remarks,AcknowOfFarmExp_Witness1Raw_ValidatorName,AcknowOfFarmExp_Witness1Raw_ValidationDate,AcknowOfFarmExp_Witness1Raw_Designation,AcknowOfFarmExp_Witness1Raw_DateWitnessed,AcknowOfFarmExp_Witness2Raw_ValidatorName,AcknowOfFarmExp_Witness2Raw_ValidationDate,AcknowOfFarmExp_Witness2Raw_Remarks,AcknowOfFarmInten_TargetArea,AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived,AcknowOfFarmInten_DateReceived,AcknowOfFarmInten_FarmerRep_Fname,AcknowOfFarmInten_FarmerRep_Mname,AcknowOfFarmInten_FarmerRep_Lname,AcknowOfFarmInten_FarmerRep_Extname,AcknowOfFarmInten_FarmerRep_Designation,AcknowOfFarmInten_Withness1_ValidatorName,AcknowOfFarmInten_Withness1_ValidationDate,AcknowOfFarmInten_Withness1_Designation,AcknowOfFarmInten_Withness1_DateWitnessed,AcknowOfFarmInten_Withness2_ValidatorName,AcknowOfFarmInten_Withness2_ValidationDate,AcknowOfFarmInten_Withness2_Remarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_FarmerRep_Fname,AcknowOfFarmRehab_FarmerRep_Mname,AcknowOfFarmRehab_FarmerRep_Lname,AcknowOfFarmRehab_FarmerRep_Extname,AcknowOfFarmRehab_FarmerRep_Designation,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_ValidationDate,AcknowOfFarmRehab_Remarks,AcknowOfProd_TypeOfProdInv,AcknowOfProd_QtyItemReceived,AcknowOfProd_DateReceived,AcknowOfProd_FarmerRep_Fname,AcknowOfProd_FarmerRep_Mname,AcknowOfProd_FarmerRep_Lname,AcknowOfProd_FarmerRep_Extname,AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,filename))
            
            insert = db.do(newMGIFO)
            
            if(insert["response"]=="error"):
                flash(f"An error occured!", "error")
                print(str(insert))
                print(sheet.name)
            else:
                flash(f"Successfully saved the new entry!", "success")
            return redirect("/mis-v4/core-mg-implementation")
        
        elif request.form.get("action") == "edit":
            print("MGI_ID raw value:", request.form.get("id"), type(int(request.form.get("id"))))

            id = int(request.form.get("id"))
            ImpUnits_RCU = request.form.get("ImpUnits_RCU")
            ImpUnits_PCU = request.form.get("ImpUnits_PCU")
            DIP_Name = request.form.get("DIP_Name")
            DIP_Commodity = request.form.get("DIP_Commodity")
            FOInfo_FOName = request.form.get("FOInfo_FOName")
            FOInfo_Address_Region = request.form.get("FOInfo_Address_Region")
            FOInfo_Address_Province = request.form.get("FOInfo_Address_Province")
            FOInfo_Address_Municipality = request.form.get("FOInfo_Address_Municipality")
            FOInfo_Address_Brgy = request.form.get("FOInfo_Address_Brgy")
            FOInfo_Address_Street = request.form.get("FOInfo_Address_Street")
            FOInfo_Head_Fname = request.form.get("FOInfo_Head_Fname")
            FOInfo_Head_Mname = request.form.get("FOInfo_Head_Mname")
            FOInfo_Head_Lname = request.form.get("FOInfo_Head_Lname")
            FOInfo_Head_Extname = request.form.get("FOInfo_Head_Extname")
            FOInfo_Head_Sex = request.form.get("FOInfo_Head_Sex")
            FOInfo_Head_Sector = request.form.get("FOInfo_Head_Sector")
            AcknowOfFarmExp_TargetArea = 0 if not request.form.get("AcknowOfFarmExp_TargetArea") else float(request.form.get("AcknowOfFarmExp_TargetArea")) 
            AcknowOfFarmExp_ItemName = request.form.get("AcknowOfFarmExp_ItemName")
            AcknowOfFarmExp_QtyReceived = 0 if not request.form.get("AcknowOfFarmExp_QtyReceived") else int(request.form.get("AcknowOfFarmExp_QtyReceived"))
            AcknowOfFarmExp_DateReceived = request.form.get("AcknowOfFarmExp_DateReceived")
            AcknowOfFarmExp_FarmerRep_Fname = request.form.get("AcknowOfFarmExp_FarmerRep_Fname")
            AcknowOfFarmExp_FarmerRep_Mname = request.form.get("AcknowOfFarmExp_FarmerRep_Mname")
            AcknowOfFarmExp_FarmerRep_Lname = request.form.get("AcknowOfFarmExp_FarmerRep_Lname")
            AcknowOfFarmExp_FarmerRep_Extname = request.form.get("AcknowOfFarmExp_FarmerRep_Extname")
            AcknowOfFarmExp_FarmerRep_Designation = request.form.get("AcknowOfFarmExp_FarmerRep_Designation")
            AcknowOfFarmExp_Witness1_ValidatorName = request.form.get("AcknowOfFarmExp_Witness1_ValidatorName")
            AcknowOfFarmExp_Witness1_ValidationDate = request.form.get("AcknowOfFarmExp_Witness1_ValidationDate")
            AcknowOfFarmExp_Witness1_Designation = request.form.get("AcknowOfFarmExp_Witness1_Designation")
            AcknowOfFarmExp_Witness1_DateWitnessed = request.form.get("AcknowOfFarmExp_Witness1_DateWitnessed")
            AcknowOfFarmExp_Witness2_ValidatorName = request.form.get("AcknowOfFarmExp_Witness2_ValidatorName")
            AcknowOfFarmExp_Witness2_ValidationDate = request.form.get("AcknowOfFarmExp_Witness2_ValidationDate")
            AcknowOfFarmExp_Witness2_Remarks = request.form.get("AcknowOfFarmExp_Witness2_Remarks")
            AcknowOfFarmExp_Witness1Raw_ValidatorName = request.form.get("AcknowOfFarmExp_Witness1Raw_ValidatorName")
            AcknowOfFarmExp_Witness1Raw_ValidationDate = request.form.get("AcknowOfFarmExp_Witness1Raw_ValidationDate")
            AcknowOfFarmExp_Witness1Raw_Designation = request.form.get("AcknowOfFarmExp_Witness1Raw_Designation")
            AcknowOfFarmExp_Witness1Raw_DateWitnessed = request.form.get("AcknowOfFarmExp_Witness1Raw_DateWitnessed")
            AcknowOfFarmExp_Witness2Raw_ValidatorName = request.form.get("AcknowOfFarmExp_Witness2Raw_ValidatorName")
            AcknowOfFarmExp_Witness2Raw_ValidationDate = request.form.get("AcknowOfFarmExp_Witness2Raw_ValidationDate")
            AcknowOfFarmExp_Witness2Raw_Remarks = request.form.get("AcknowOfFarmExp_Witness2Raw_Remarks")
            AcknowOfFarmInten_TargetArea = 0 if not request.form.get("AcknowOfFarmInten_TargetArea") else float(request.form.get("AcknowOfFarmInten_TargetArea"))
            AcknowOfFarmInten_ItemName = request.form.get("AcknowOfFarmInten_ItemName")
            AcknowOfFarmInten_QtyReceived = 0 if not request.form.get("AcknowOfFarmInten_QtyReceived") else int(request.form.get("AcknowOfFarmInten_QtyReceived"))
            AcknowOfFarmInten_DateReceived = request.form.get("AcknowOfFarmInten_DateReceived")
            AcknowOfFarmInten_FarmerRep_Fname = request.form.get("AcknowOfFarmInten_FarmerRep_Fname")
            AcknowOfFarmInten_FarmerRep_Mname = request.form.get("AcknowOfFarmInten_FarmerRep_Mname")
            AcknowOfFarmInten_FarmerRep_Lname = request.form.get("AcknowOfFarmInten_FarmerRep_Lname")
            AcknowOfFarmInten_FarmerRep_Extname = request.form.get("AcknowOfFarmInten_FarmerRep_Extname")
            AcknowOfFarmInten_FarmerRep_Designation = request.form.get("AcknowOfFarmInten_FarmerRep_Designation")
            AcknowOfFarmInten_Withness1_ValidatorName = request.form.get("AcknowOfFarmInten_Withness1_ValidatorName")
            AcknowOfFarmInten_Withness1_ValidationDate = request.form.get("AcknowOfFarmInten_Withness1_ValidationDate")
            AcknowOfFarmInten_Withness1_Designation = request.form.get("AcknowOfFarmInten_Withness1_Designation")
            AcknowOfFarmInten_Withness1_DateWitnessed = request.form.get("AcknowOfFarmInten_Withness1_DateWitnessed")
            AcknowOfFarmInten_Withness2_ValidatorName = request.form.get("AcknowOfFarmInten_Withness2_ValidatorName")
            AcknowOfFarmInten_Withness2_ValidationDate = request.form.get("AcknowOfFarmInten_Withness2_ValidationDate")
            AcknowOfFarmInten_Withness2_Remarks = request.form.get("AcknowOfFarmInten_Withness2_Remarks")
            AcknowOfFarmRehab_TargetArea = 0 if not request.form.get("AcknowOfFarmRehab_TargetArea") else float(request.form.get("AcknowOfFarmRehab_TargetArea"))
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw = 0 if not request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw") else int(request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw"))
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw = 0 if not request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw") else int(request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw"))
            AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer = 0 if not request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer") else int(request.form.get("AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer"))
            AcknowOfFarmRehab_DateReceived = request.form.get("AcknowOfFarmRehab_DateReceived")
            AcknowOfFarmRehab_FarmerRep_Fname = request.form.get("AcknowOfFarmRehab_FarmerRep_Fname")
            AcknowOfFarmRehab_FarmerRep_Mname = request.form.get("AcknowOfFarmRehab_FarmerRep_Mname")
            AcknowOfFarmRehab_FarmerRep_Lname = request.form.get("AcknowOfFarmRehab_FarmerRep_Lname")
            AcknowOfFarmRehab_FarmerRep_Extname = request.form.get("AcknowOfFarmRehab_FarmerRep_Extname")
            AcknowOfFarmRehab_FarmerRep_Designation = request.form.get("AcknowOfFarmRehab_FarmerRep_Designation")
            AcknowOfFarmRehab_ValidatorName = request.form.get("AcknowOfFarmRehab_ValidatorName")
            AcknowOfFarmRehab_ValidationDate = request.form.get("AcknowOfFarmRehab_ValidationDate")
            AcknowOfFarmRehab_Remarks = request.form.get("AcknowOfFarmRehab_Remarks")
            AcknowOfProd_TypeOfProdInv = request.form.get("AcknowOfProd_TypeOfProdInv")
            AcknowOfProd_QtyItemReceived = 0 if not request.form.get("AcknowOfProd_QtyItemReceived") else int(request.form.get("AcknowOfProd_QtyItemReceived"))
            AcknowOfProd_DateReceived = request.form.get("AcknowOfProd_DateReceived")
            AcknowOfProd_FarmerRep_Fname = request.form.get("AcknowOfProd_FarmerRep_Fname")
            AcknowOfProd_FarmerRep_Mname = request.form.get("AcknowOfProd_FarmerRep_Mname")
            AcknowOfProd_FarmerRep_Lname = request.form.get("AcknowOfProd_FarmerRep_Lname")
            AcknowOfProd_FarmerRep_Extname = request.form.get("AcknowOfProd_FarmerRep_Extname")
            AcknowOfProd_FarmerRep_Designation = request.form.get("AcknowOfProd_FarmerRep_Designation")
            AcknowOfProd_ValidatorName = request.form.get("AcknowOfProd_ValidatorName")
            AcknowOfProd_ValidationDate = request.form.get("AcknowOfProd_ValidationDate")
            AcknowOfProd_Remarks = request.form.get("AcknowOfProd_Remarks")
            filename = ''

            updateMGIFO = ("UPDATE mg_implementation_fo SET ImpUnits_RCU = '{}', ImpUnits_PCU = '{}', DIP_Name = '{}', DIP_Commodity = '{}', FOInfo_FOName = '{}', FOInfo_Address_Region = '{}', FOInfo_Address_Province = '{}', FOInfo_Address_Municipality = '{}', FOInfo_Address_Brgy = '{}', FOInfo_Address_Street = '{}', FOInfo_Head_Fname = '{}', FOInfo_Head_Mname = '{}', FOInfo_Head_Lname = '{}', FOInfo_Head_Extname = '{}', FOInfo_Head_Sex = '{}', FOInfo_Head_Sector = '{}', AcknowOfFarmExp_TargetArea = '{}', AcknowOfFarmExp_ItemName = '{}', AcknowOfFarmExp_QtyReceived = '{}', AcknowOfFarmExp_DateReceived = '{}', AcknowOfFarmExp_FarmerRep_Fname = '{}', AcknowOfFarmExp_FarmerRep_Mname = '{}', AcknowOfFarmExp_FarmerRep_Lname = '{}', AcknowOfFarmExp_FarmerRep_Extname = '{}', AcknowOfFarmExp_FarmerRep_Designation = '{}', AcknowOfFarmExp_Witness1_ValidatorName = '{}', AcknowOfFarmExp_Witness1_ValidationDate = '{}', AcknowOfFarmExp_Witness1_Designation = '{}', AcknowOfFarmExp_Witness1_DateWitnessed = '{}', AcknowOfFarmExp_Witness2_ValidatorName = '{}', AcknowOfFarmExp_Witness2_ValidationDate = '{}', AcknowOfFarmExp_Witness2_Remarks = '{}', AcknowOfFarmExp_Witness1Raw_ValidatorName = '{}', AcknowOfFarmExp_Witness1Raw_ValidationDate = '{}', AcknowOfFarmExp_Witness1Raw_Designation = '{}', AcknowOfFarmExp_Witness1Raw_DateWitnessed = '{}', AcknowOfFarmExp_Witness2Raw_ValidatorName = '{}', AcknowOfFarmExp_Witness2Raw_ValidationDate = '{}', AcknowOfFarmExp_Witness2Raw_Remarks = '{}', AcknowOfFarmInten_TargetArea = '{}', AcknowOfFarmInten_ItemName = '{}', AcknowOfFarmInten_QtyReceived = '{}', AcknowOfFarmInten_DateReceived = '{}', AcknowOfFarmInten_FarmerRep_Fname = '{}', AcknowOfFarmInten_FarmerRep_Mname = '{}', AcknowOfFarmInten_FarmerRep_Lname = '{}', AcknowOfFarmInten_FarmerRep_Extname = '{}', AcknowOfFarmInten_FarmerRep_Designation = '{}', AcknowOfFarmInten_Withness1_ValidatorName = '{}', AcknowOfFarmInten_Withness1_ValidationDate = '{}', AcknowOfFarmInten_Withness1_Designation = '{}', AcknowOfFarmInten_Withness1_DateWitnessed = '{}', AcknowOfFarmInten_Withness2_ValidatorName = '{}', AcknowOfFarmInten_Withness2_ValidationDate = '{}', AcknowOfFarmInten_Withness2_Remarks = '{}', AcknowOfFarmRehab_TargetArea = '{}', AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw = '{}', AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw = '{}', AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer = '{}', AcknowOfFarmRehab_DateReceived = '{}', AcknowOfFarmRehab_FarmerRep_Fname = '{}', AcknowOfFarmRehab_FarmerRep_Mname = '{}', AcknowOfFarmRehab_FarmerRep_Lname = '{}', AcknowOfFarmRehab_FarmerRep_Extname = '{}', AcknowOfFarmRehab_FarmerRep_Designation = '{}', AcknowOfFarmRehab_ValidatorName = '{}', AcknowOfFarmRehab_ValidationDate = '{}', AcknowOfFarmRehab_Remarks = '{}', AcknowOfProd_TypeOfProdInv = '{}', AcknowOfProd_QtyItemReceived = '{}', AcknowOfProd_DateReceived = '{}', AcknowOfProd_FarmerRep_Fname = '{}', AcknowOfProd_FarmerRep_Mname = '{}', AcknowOfProd_FarmerRep_Lname = '{}', AcknowOfProd_FarmerRep_Extname = '{}', AcknowOfProd_FarmerRep_Designation = '{}', AcknowOfProd_ValidatorName = '{}', AcknowOfProd_ValidationDate = '{}', AcknowOfProd_Remarks = '{}', DateModified = NOW() WHERE MGI_ID = {}".
                           format(ImpUnits_RCU, ImpUnits_PCU, DIP_Name, DIP_Commodity, FOInfo_FOName, FOInfo_Address_Region, FOInfo_Address_Province, FOInfo_Address_Municipality, FOInfo_Address_Brgy, FOInfo_Address_Street, FOInfo_Head_Fname, FOInfo_Head_Mname, FOInfo_Head_Lname, FOInfo_Head_Extname, FOInfo_Head_Sex, FOInfo_Head_Sector, AcknowOfFarmExp_TargetArea, AcknowOfFarmExp_ItemName, AcknowOfFarmExp_QtyReceived, AcknowOfFarmExp_DateReceived, AcknowOfFarmExp_FarmerRep_Fname, AcknowOfFarmExp_FarmerRep_Mname, AcknowOfFarmExp_FarmerRep_Lname, AcknowOfFarmExp_FarmerRep_Extname, AcknowOfFarmExp_FarmerRep_Designation, AcknowOfFarmExp_Witness1_ValidatorName, AcknowOfFarmExp_Witness1_ValidationDate, AcknowOfFarmExp_Witness1_Designation, AcknowOfFarmExp_Witness1_DateWitnessed, AcknowOfFarmExp_Witness2_ValidatorName, AcknowOfFarmExp_Witness2_ValidationDate, AcknowOfFarmExp_Witness2_Remarks, AcknowOfFarmExp_Witness1Raw_ValidatorName, AcknowOfFarmExp_Witness1Raw_ValidationDate, AcknowOfFarmExp_Witness1Raw_Designation, AcknowOfFarmExp_Witness1Raw_DateWitnessed, AcknowOfFarmExp_Witness2Raw_ValidatorName, AcknowOfFarmExp_Witness2Raw_ValidationDate, AcknowOfFarmExp_Witness2Raw_Remarks, AcknowOfFarmInten_TargetArea, AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived, AcknowOfFarmInten_DateReceived, AcknowOfFarmInten_FarmerRep_Fname, AcknowOfFarmInten_FarmerRep_Mname, AcknowOfFarmInten_FarmerRep_Lname, AcknowOfFarmInten_FarmerRep_Extname, AcknowOfFarmInten_FarmerRep_Designation, AcknowOfFarmInten_Withness1_ValidatorName, AcknowOfFarmInten_Withness1_ValidationDate, AcknowOfFarmInten_Withness1_Designation, AcknowOfFarmInten_Withness1_DateWitnessed, AcknowOfFarmInten_Withness2_ValidatorName, AcknowOfFarmInten_Withness2_ValidationDate, AcknowOfFarmInten_Withness2_Remarks, AcknowOfFarmRehab_TargetArea, AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw, AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw, AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer, AcknowOfFarmRehab_DateReceived, AcknowOfFarmRehab_FarmerRep_Fname, AcknowOfFarmRehab_FarmerRep_Mname, AcknowOfFarmRehab_FarmerRep_Lname, AcknowOfFarmRehab_FarmerRep_Extname, AcknowOfFarmRehab_FarmerRep_Designation, AcknowOfFarmRehab_ValidatorName, AcknowOfFarmRehab_ValidationDate, AcknowOfFarmRehab_Remarks, AcknowOfProd_TypeOfProdInv, AcknowOfProd_QtyItemReceived, AcknowOfProd_DateReceived, AcknowOfProd_FarmerRep_Fname, AcknowOfProd_FarmerRep_Mname, AcknowOfProd_FarmerRep_Lname, AcknowOfProd_FarmerRep_Extname, AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,id))
            
            update = db.do(updateMGIFO)
            # print("ImpUnits_RCU:", ImpUnits_RCU, type(ImpUnits_RCU))
            # print("ImpUnits_PCU:", ImpUnits_PCU, type(ImpUnits_PCU))
            # print("id:", id, type(id))
            # print("SQL : ", updateMGIFO)
            
            if (update["response"]=="error"):
                flash(f"An error occured!", "error")
                # print(str(update))
            else:
                flash(f"Successfully saved the edited data!", "success")
            return redirect("/mis-v4/core-mg-implementation")
        
    @app.route('/mgimplementation_shf_form/submit', methods=['POST'])
    def submit_shf_form():
        if not is_on_session():
            return redirect("/login")
        
        if request.form.get("action") == "addnew":
            UploadedBy = session["USER_DATA"][0]['id']
            ImpUnits_RCU = request.form.get("ImpUnits_RCU")
            ImpUnits_PCU = request.form.get("ImpUnits_PCU")
            DIP_Name = request.form.get("DIP_Name")
            DIP_Commodity = request.form.get("DIP_Commodity")
            MembershipToCOOP = request.form.get("MembershipToCOOP")
            FarmerInfo_Fname = request.form.get("FarmerInfo_Fname")
            FarmerInfo_Mname = request.form.get("FarmerInfo_Mname")
            FarmerInfo_Lname = request.form.get("FarmerInfo_Lname")
            FarmerInfo_ExtName = request.form.get("FarmerInfo_ExtName")
            FarmerInfo_DOB = request.form.get("FarmerInfo_DOB")
            FarmerInfo_Sex = request.form.get("FarmerInfo_Sex")
            FarmerInfo_Sector_isPWD = ( 1 if request.form.get("FarmerInfo_Sector_isPWD") else 0 )
            FarmerInfo_Sector_isYouth = ( 1 if request.form.get("FarmerInfo_Sector_isYouth") else 0 )
            FarmerInfo_Sector_isIP = ( 1 if request.form.get("FarmerInfo_Sector_isIP") else 0 )
            FarmerInfo_Sector_isSC = ( 1 if request.form.get("FarmerInfo_Sector_isSC") else 0 )
            FarmerInfo_address_Region = request.form.get("FarmerInfo_address_Region")
            FarmerInfo_address_Province = request.form.get("FarmerInfo_address_Province")
            FarmerInfo_address_municipality = request.form.get("FarmerInfo_address_municipality")
            FarmerInfo_address_brgy = request.form.get("FarmerInfo_address_brgy")
            FarmerInfo_address_street = request.form.get("FarmerInfo_address_street")
            FarmerInfo_farmLoc_long = request.form.get("FarmerInfo_farmLoc_long")
            FarmerInfo_farmLoc_lat = request.form.get("FarmerInfo_farmLoc_lat")
            Acknowledgement_NotPartOfMG = ( 1 if request.form.get("Acknowledgement_NotPartOfMG") else 0 )
            Acknowledgement_TargetArea = request.form.get("Acknowledgement_TargetArea")
            Acknowledgement_Name = request.form.get("Acknowledgement_Name")
            Acknowledgement_Qty = request.form.get("Acknowledgement_Qty")
            Acknowledgement_DateReceived = request.form.get("Acknowledgement_DateReceived")
            Acknowledgement_ValidatorName = request.form.get("Acknowledgement_ValidatorName")
            Acknowledgement_DateValidation = request.form.get("Acknowledgement_DateValidation")
            Acknowledgement_Remarks = request.form.get("Acknowledgement_Remarks")
            PlantedArea_GAP_NotPartOfMG = ( 1 if request.form.get("PlantedArea_GAP_NotPartOfMG") else 0 )
            PlantedArea_GAP_PlantingTimeline = request.form.get("PlantedArea_GAP_PlantingTimeline")
            PlantedArea_GAP_TotalAreaPlanted = request.form.get("PlantedArea_GAP_TotalAreaPlanted")
            PlantedArea_GAP_NoOfSeedlingsPlanted = request.form.get("PlantedArea_GAP_NoOfSeedlingsPlanted")
            PlantedArea_GAP_NoOfSeedlingsNotPlanted = request.form.get("PlantedArea_GAP_NoOfSeedlingsNotPlanted")
            PlantedArea_GAP_MortalityRate = request.form.get("PlantedArea_GAP_MortalityRate")
            PlantedArea_GAP_PACE_isPlantingDistance = ( 1 if request.form.get("PlantedArea_GAP_PACE_isPlantingDistance") else 0 )
            PlantedArea_GAP_PACE_isDiggingHoles = ( 1 if request.form.get("PlantedArea_GAP_PACE_isDiggingHoles") else 0 )
            PlantedArea_GAP_PACE_isSeparateTopSoil = ( 1 if request.form.get("PlantedArea_GAP_PACE_isSeparateTopSoil") else 0 )
            PlantedArea_GAP_PACE_isRemovingPlasticBag = ( 1 if request.form.get("PlantedArea_GAP_PACE_isRemovingPlasticBag") else 0 )
            PlantedArea_GAP_PACE_isApplyingMulching = ( 1 if request.form.get("PlantedArea_GAP_PACE_isApplyingMulching") else 0 )
            PlantedArea_GAP_PACE_isShadingEstablishment = ( 1 if request.form.get("PlantedArea_GAP_PACE_isShadingEstablishment") else 0 )
            
               
            PlantedArea_SOI_isOwn = ( 1 if request.form.get("PlantedArea_SOI")=="own" else 0 )
            PlantedArea_SOI_isGovAgency = ( 1 if request.form.get("PlantedArea_SOI")=="gov" else 0 )
            PlantedArea_SOI_isLoan = ( 1 if request.form.get("PlantedArea_SOI")=="loan" else 0 )
            PlantedArea_SOI_Others = ( request.form.get('PlantedArea_SOI_others') if request.form.get("PlantedArea_SOI")=="others" else 0 )
            
            PlantedArea_GAP_SANM_isBasal = ( 1 if request.form.get("PlantedArea_GAP_SANM_isBasal") else 0 )
            PlantedArea_GAP_SANM_isFoliar = ( 1 if request.form.get("PlantedArea_GAP_SANM_isFoliar") else 0 )
            PlantedArea_GAP_SANM_isPesticide = ( 1 if request.form.get("PlantedArea_GAP_SANM_isPesticide") else 0 )
            PlantedArea_GAP_SANM_isPlantBased = ( 1 if request.form.get("PlantedArea_GAP_SANM_isPlantBased") else 0 )
            
            PlantedArea_ValidatorName = request.form.get("PlantedArea_ValidatorName")
            PlantedArea_DateValidation = request.form.get("PlantedArea_DateValidation")
            PlantedArea_ValidationRemarks = request.form.get("PlantedArea_ValidationRemarks")
            
            AcknowOfFarmIntensification_NotPartOfMG = ( 1 if request.form.get("AcknowOfFarmIntensification_NotPartOfMG") else 0 )
            AcknowOfFarmIntensification_TargetArea = request.form.get("AcknowOfFarmIntensification_TargetArea")
            AcknowOfFarmIntensification_Name = request.form.get("AcknowOfFarmIntensification_Name")
            AcknowOfFarmIntensification_Qty = request.form.get("AcknowOfFarmIntensification_Qty")
            AcknowOfFarmIntensification_DateReceived = request.form.get("AcknowOfFarmIntensification_DateReceived")
            AcknowOfFarmIntensification_ValidatorName = request.form.get("AcknowOfFarmIntensification_ValidatorName")
            AcknowOfFarmIntensification_DateValidation = request.form.get("AcknowOfFarmIntensification_DateValidation")
            AcknowOfFarmIntensification_ValidationRemarks = request.form.get("AcknowOfFarmIntensification_ValidationRemarks")
            PlantedAreaForIntens_NotPartOfMG = ( 1 if request.form.get("PlantedAreaForIntens_NotPartOfMG") else 0 )
            PlantedAreaForIntens_PlantingTimeline = request.form.get("PlantedAreaForIntens_PlantingTimeline")
            PlantedAreaForIntens_TotalAreaPlanted = request.form.get("PlantedAreaForIntens_TotalAreaPlanted")
            PlantedAreaForIntens_NoOfSeedlingsPlanted = request.form.get("PlantedAreaForIntens_NoOfSeedlingsPlanted")
            PlantedAreaForIntens_NoOfSeedlingsNotPlanted = request.form.get("PlantedAreaForIntens_NoOfSeedlingsNotPlanted")
            PlantedAreaForIntens_PlantingDistance = request.form.get("PlantedAreaForIntens_PlantingDistance")
            PlantedAreaForIntens_HoleDepth = request.form.get("PlantedAreaForIntens_HoleDepth")
            PlantedAreaForIntens_MortalityRate = request.form.get("PlantedAreaForIntens_MortalityRate")
            PlantedAreaForIntens_ValidatedBy = request.form.get("PlantedAreaForIntens_ValidatedBy")
            PlantedAreaForIntens_DateValidation = request.form.get("PlantedAreaForIntens_DateValidation")
            PlantedAreaForIntens_ValidationRemarks = request.form.get("PlantedAreaForIntens_ValidationRemarks")
            AcknowOfFarmRehab_NotPartOfMG = ( 1 if request.form.get("AcknowOfFarmRehab_NotPartOfMG") else 0 )
            AcknowOfFarmRehab_TargetArea = request.form.get("AcknowOfFarmRehab_TargetArea")
            AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear")
            AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw")
            AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife")
            AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape")
            AcknowOfFarmRehab_QtyRehabToolsReceived_Others = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_Others")
            AcknowOfFarmRehab_DateReceived = request.form.get("AcknowOfFarmRehab_DateReceived")
            AcknowOfFarmRehab_ValidatorName = request.form.get("AcknowOfFarmRehab_ValidatorName")
            AcknowOfFarmRehab_DateValidation = request.form.get("AcknowOfFarmRehab_DateValidation")
            AcknowOfFarmRehab_ValidationRemarks = request.form.get("AcknowOfFarmRehab_ValidationRemarks")
            RehabFarmArea_GAP_NotParOfMG = ( 1 if request.form.get("RehabFarmArea_GAP_NotParOfMG") else 0 )
            RehabFarmArea_GAP_RehabTimeline = request.form.get("RehabFarmArea_GAP_RehabTimeline")
            RehabFarmArea_GAP_TotalAreaRehab = request.form.get("RehabFarmArea_GAP_TotalAreaRehab")
            RehabFarmArea_GAP_NoOfTreesRehab = request.form.get("RehabFarmArea_GAP_NoOfTreesRehab")
            RehabFarmArea_GAP_isShadingMaintenance = request.form.get("RehabFarmArea_GAP_isShadingMaintenance")
            RehabFarmArea_GAP_isTipPruning = request.form.get("RehabFarmArea_GAP_isTipPruning")
            RehabFarmArea_GAP_isShapePruning = request.form.get("RehabFarmArea_GAP_isShapePruning")
            RehabFarmArea_GAP_isAccessPruning = request.form.get("RehabFarmArea_GAP_isAccessPruning")
            RehabFarmArea_GAP_isSanitaryPruning = request.form.get("RehabFarmArea_GAP_isSanitaryPruning")
            RehabFarmArea_GAP_isMaintenancePruning = request.form.get("RehabFarmArea_GAP_isMaintenancePruning")
            RehabFarmArea_GAP_isUsingApprTools = request.form.get("RehabFarmArea_GAP_isUsingApprTools")
            RehabFarmArea_GAP_isFarmRecordKeeping = request.form.get("RehabFarmArea_GAP_isFarmRecordKeeping")
            RehabFarmArea_GAP_isChupon = request.form.get("RehabFarmArea_GAP_isChupon")
            RehabFarmArea_GAP_isFoliarFertilizer = request.form.get("RehabFarmArea_GAP_isFoliarFertilizer")
            RehabFarmArea_GAP_isSealing = request.form.get("RehabFarmArea_GAP_isSealing")
            RehabFarmArea_GAP_isObservance = request.form.get("RehabFarmArea_GAP_isObservance")
            RehabFarmArea_GAP_isGrafting = request.form.get("RehabFarmArea_GAP_isGrafting")
            RehabFarmArea_GAP_ValidatorName = request.form.get("RehabFarmArea_GAP_ValidatorName")
            RehabFarmArea_GAP_DateValidation = request.form.get("RehabFarmArea_GAP_DateValidation")
            RehabFarmArea_GAP_ValidationRemarks = request.form.get("RehabFarmArea_GAP_ValidationRemarks")
            AccessToProdInvestment_NotPartOfMG = ( 1 if request.form.get("AccessToProdInvestment_NotPartOfMG") else 0 )
            AccessToProdInvestment_TypeOfProdInvestments = request.form.get("AccessToProdInvestment_TypeOfProdInvestments")
            AccessToProdInvestment_ValidatorName = request.form.get("AccessToProdInvestment_ValidatorName")
            AccessToProdInvestment_DateValidation = request.form.get("AccessToProdInvestment_DateValidation")
            AccessToProdInvestment_ValidationRemarks = request.form.get("AccessToProdInvestment_ValidationRemarks")
            filename = ''
            
            
            newMGISHF = ("INSERT INTO mg_implementation_shf (UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,MembershipToCOOP,FarmerInfo_Fname,FarmerInfo_Mname,FarmerInfo_Lname,FarmerInfo_ExtName,FarmerInfo_DOB,FarmerInfo_Sex,FarmerInfo_Sector_isPWD,FarmerInfo_Sector_isYouth,FarmerInfo_Sector_isIP,FarmerInfo_Sector_isSC,FarmerInfo_address_Region,FarmerInfo_address_Province,FarmerInfo_address_municipality,FarmerInfo_address_brgy,FarmerInfo_address_street,FarmerInfo_farmLoc_long,FarmerInfo_farmLoc_lat,Acknowledgement_NotPartOfMG,Acknowledgement_TargetArea,Acknowledgement_Name,Acknowledgement_Qty,Acknowledgement_DateReceived,Acknowledgement_ValidatorName,Acknowledgement_DateValidation,Acknowledgement_Remarks,PlantedArea_GAP_NotPartOfMG,PlantedArea_GAP_PlantingTimeline,PlantedArea_GAP_TotalAreaPlanted,PlantedArea_GAP_NoOfSeedlingsPlanted,PlantedArea_GAP_NoOfSeedlingsNotPlanted,PlantedArea_GAP_MortalityRate,PlantedArea_GAP_PACE_isPlantingDistance,PlantedArea_GAP_PACE_isDiggingHoles,PlantedArea_GAP_PACE_isSeparateTopSoil,PlantedArea_GAP_PACE_isRemovingPlasticBag,PlantedArea_GAP_PACE_isApplyingMulching,PlantedArea_GAP_PACE_isShadingEstablishment,PlantedArea_SOI_isOwn,PlantedArea_SOI_isGovAgency,PlantedArea_SOI_isLoan,PlantedArea_SOI_Others,PlantedArea_GAP_SANM_isBasal,PlantedArea_GAP_SANM_isFoliar,PlantedArea_GAP_SANM_isPesticide,PlantedArea_GAP_SANM_isPlantBased,PlantedArea_ValidatorName,PlantedArea_DateValidation,PlantedArea_ValidationRemarks,AcknowOfFarmIntensification_NotPartOfMG,AcknowOfFarmIntensification_TargetArea,AcknowOfFarmIntensification_Name,AcknowOfFarmIntensification_Qty,AcknowOfFarmIntensification_DateReceived,AcknowOfFarmIntensification_ValidatorName,AcknowOfFarmIntensification_DateValidation,AcknowOfFarmIntensification_ValidationRemarks,PlantedAreaForIntens_NotPartOfMG,PlantedAreaForIntens_PlantingTimeline,PlantedAreaForIntens_TotalAreaPlanted,PlantedAreaForIntens_NoOfSeedlingsPlanted,PlantedAreaForIntens_NoOfSeedlingsNotPlanted,PlantedAreaForIntens_PlantingDistance,PlantedAreaForIntens_HoleDepth,PlantedAreaForIntens_MortalityRate,PlantedAreaForIntens_ValidatedBy,PlantedAreaForIntens_DateValidation,PlantedAreaForIntens_ValidationRemarks,AcknowOfFarmRehab_NotPartOfMG,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw,AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife,AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape,AcknowOfFarmRehab_QtyRehabToolsReceived_Others,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_DateValidation,AcknowOfFarmRehab_ValidationRemarks,RehabFarmArea_GAP_NotParOfMG,RehabFarmArea_GAP_RehabTimeline,RehabFarmArea_GAP_TotalAreaRehab,RehabFarmArea_GAP_NoOfTreesRehab,RehabFarmArea_GAP_isShadingMaintenance,RehabFarmArea_GAP_isTipPruning,RehabFarmArea_GAP_isShapePruning,RehabFarmArea_GAP_isAccessPruning,RehabFarmArea_GAP_isSanitaryPruning,RehabFarmArea_GAP_isMaintenancePruning,RehabFarmArea_GAP_isUsingApprTools,RehabFarmArea_GAP_isFarmRecordKeeping,RehabFarmArea_GAP_isChupon,RehabFarmArea_GAP_isFoliarFertilizer,RehabFarmArea_GAP_isSealing,RehabFarmArea_GAP_isObservance,RehabFarmArea_GAP_isGrafting,RehabFarmArea_GAP_ValidatorName,RehabFarmArea_GAP_DateValidation,RehabFarmArea_GAP_ValidationRemarks,AccessToProdInvestment_NotPartOfMG,AccessToProdInvestment_TypeOfProdInvestments,AccessToProdInvestment_ValidatorName,AccessToProdInvestment_DateValidation,AccessToProdInvestment_ValidationRemarks,filename) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".
            format(UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,MembershipToCOOP,FarmerInfo_Fname,FarmerInfo_Mname,FarmerInfo_Lname,FarmerInfo_ExtName,FarmerInfo_DOB,FarmerInfo_Sex,FarmerInfo_Sector_isPWD,FarmerInfo_Sector_isYouth,FarmerInfo_Sector_isIP,FarmerInfo_Sector_isSC,FarmerInfo_address_Region,FarmerInfo_address_Province,FarmerInfo_address_municipality,FarmerInfo_address_brgy,FarmerInfo_address_street,FarmerInfo_farmLoc_long,FarmerInfo_farmLoc_lat,Acknowledgement_NotPartOfMG,Acknowledgement_TargetArea,Acknowledgement_Name,Acknowledgement_Qty,Acknowledgement_DateReceived,Acknowledgement_ValidatorName,Acknowledgement_DateValidation,Acknowledgement_Remarks,PlantedArea_GAP_NotPartOfMG,PlantedArea_GAP_PlantingTimeline,PlantedArea_GAP_TotalAreaPlanted,PlantedArea_GAP_NoOfSeedlingsPlanted,PlantedArea_GAP_NoOfSeedlingsNotPlanted,PlantedArea_GAP_MortalityRate,PlantedArea_GAP_PACE_isPlantingDistance,PlantedArea_GAP_PACE_isDiggingHoles,PlantedArea_GAP_PACE_isSeparateTopSoil,PlantedArea_GAP_PACE_isRemovingPlasticBag,PlantedArea_GAP_PACE_isApplyingMulching,PlantedArea_GAP_PACE_isShadingEstablishment,PlantedArea_SOI_isOwn,PlantedArea_SOI_isGovAgency,PlantedArea_SOI_isLoan,PlantedArea_SOI_Others,PlantedArea_GAP_SANM_isBasal,PlantedArea_GAP_SANM_isFoliar,PlantedArea_GAP_SANM_isPesticide,PlantedArea_GAP_SANM_isPlantBased,PlantedArea_ValidatorName,PlantedArea_DateValidation,PlantedArea_ValidationRemarks,AcknowOfFarmIntensification_NotPartOfMG,AcknowOfFarmIntensification_TargetArea,AcknowOfFarmIntensification_Name,AcknowOfFarmIntensification_Qty,AcknowOfFarmIntensification_DateReceived,AcknowOfFarmIntensification_ValidatorName,AcknowOfFarmIntensification_DateValidation,AcknowOfFarmIntensification_ValidationRemarks,PlantedAreaForIntens_NotPartOfMG,PlantedAreaForIntens_PlantingTimeline,PlantedAreaForIntens_TotalAreaPlanted,PlantedAreaForIntens_NoOfSeedlingsPlanted,PlantedAreaForIntens_NoOfSeedlingsNotPlanted,PlantedAreaForIntens_PlantingDistance,PlantedAreaForIntens_HoleDepth,PlantedAreaForIntens_MortalityRate,PlantedAreaForIntens_ValidatedBy,PlantedAreaForIntens_DateValidation,PlantedAreaForIntens_ValidationRemarks,AcknowOfFarmRehab_NotPartOfMG,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw,AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife,AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape,AcknowOfFarmRehab_QtyRehabToolsReceived_Others,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_DateValidation,AcknowOfFarmRehab_ValidationRemarks,RehabFarmArea_GAP_NotParOfMG,RehabFarmArea_GAP_RehabTimeline,RehabFarmArea_GAP_TotalAreaRehab,RehabFarmArea_GAP_NoOfTreesRehab,RehabFarmArea_GAP_isShadingMaintenance,RehabFarmArea_GAP_isTipPruning,RehabFarmArea_GAP_isShapePruning,RehabFarmArea_GAP_isAccessPruning,RehabFarmArea_GAP_isSanitaryPruning,RehabFarmArea_GAP_isMaintenancePruning,RehabFarmArea_GAP_isUsingApprTools,RehabFarmArea_GAP_isFarmRecordKeeping,RehabFarmArea_GAP_isChupon,RehabFarmArea_GAP_isFoliarFertilizer,RehabFarmArea_GAP_isSealing,RehabFarmArea_GAP_isObservance,RehabFarmArea_GAP_isGrafting,RehabFarmArea_GAP_ValidatorName,RehabFarmArea_GAP_DateValidation,RehabFarmArea_GAP_ValidationRemarks,AccessToProdInvestment_NotPartOfMG,AccessToProdInvestment_TypeOfProdInvestments,AccessToProdInvestment_ValidatorName,AccessToProdInvestment_DateValidation,AccessToProdInvestment_ValidationRemarks,filename))
            
            insert = db.do(newMGISHF)
            
            if(insert["response"]=="error"):
                flash(f"An error occured!", "error")
                print(str(insert))
                # print(sheet.name)
            else:
                flash(f"Successfully saved the new entry!", "success")
            return redirect("/mis-v4/core-mg-implementation")
        
        elif request.form.get("action") == "edit":
            print("MGI_ID raw value:", request.form.get("id"), type(int(request.form.get("id"))))

            id = int(request.form.get("id"))
            ImpUnits_RCU = request.form.get("ImpUnits_RCU")
            ImpUnits_PCU = request.form.get("ImpUnits_PCU")
            DIP_Name = request.form.get("DIP_Name")
            DIP_Commodity = request.form.get("DIP_Commodity")
            MembershipToCOOP = request.form.get("MembershipToCOOP")
            FarmerInfo_Fname = request.form.get("FarmerInfo_Fname")
            FarmerInfo_Mname = request.form.get("FarmerInfo_Mname")
            FarmerInfo_Lname = request.form.get("FarmerInfo_Lname")
            FarmerInfo_ExtName = request.form.get("FarmerInfo_ExtName")
            FarmerInfo_DOB = request.form.get("FarmerInfo_DOB")
            FarmerInfo_Sex = request.form.get("FarmerInfo_Sex")
            FarmerInfo_Sector_isPWD = ( 1 if request.form.get("FarmerInfo_Sector_isPWD") else 0 )
            FarmerInfo_Sector_isYouth = ( 1 if request.form.get("FarmerInfo_Sector_isYouth") else 0 )
            FarmerInfo_Sector_isIP = ( 1 if request.form.get("FarmerInfo_Sector_isIP") else 0 )
            FarmerInfo_Sector_isSC = ( 1 if request.form.get("FarmerInfo_Sector_isSC") else 0 )
            FarmerInfo_address_Region = request.form.get("FarmerInfo_address_Region")
            FarmerInfo_address_Province = request.form.get("FarmerInfo_address_Province")
            FarmerInfo_address_municipality = request.form.get("FarmerInfo_address_municipality")
            FarmerInfo_address_brgy = request.form.get("FarmerInfo_address_brgy")
            FarmerInfo_address_street = request.form.get("FarmerInfo_address_street")
            FarmerInfo_farmLoc_long = request.form.get("FarmerInfo_farmLoc_long")
            FarmerInfo_farmLoc_lat = request.form.get("FarmerInfo_farmLoc_lat")
            Acknowledgement_NotPartOfMG = ( 1 if request.form.get("Acknowledgement_NotPartOfMG") else 0 )
            Acknowledgement_TargetArea = request.form.get("Acknowledgement_TargetArea")
            Acknowledgement_Name = request.form.get("Acknowledgement_Name")
            Acknowledgement_Qty = request.form.get("Acknowledgement_Qty")
            Acknowledgement_DateReceived = request.form.get("Acknowledgement_DateReceived")
            Acknowledgement_ValidatorName = request.form.get("Acknowledgement_ValidatorName")
            Acknowledgement_DateValidation = request.form.get("Acknowledgement_DateValidation")
            Acknowledgement_Remarks = request.form.get("Acknowledgement_Remarks")
            PlantedArea_GAP_NotPartOfMG = ( 1 if request.form.get("PlantedArea_GAP_NotPartOfMG") else 0 )
            PlantedArea_GAP_PlantingTimeline = request.form.get("PlantedArea_GAP_PlantingTimeline")
            PlantedArea_GAP_TotalAreaPlanted = request.form.get("PlantedArea_GAP_TotalAreaPlanted")
            PlantedArea_GAP_NoOfSeedlingsPlanted = request.form.get("PlantedArea_GAP_NoOfSeedlingsPlanted")
            PlantedArea_GAP_NoOfSeedlingsNotPlanted = request.form.get("PlantedArea_GAP_NoOfSeedlingsNotPlanted")
            PlantedArea_GAP_MortalityRate = request.form.get("PlantedArea_GAP_MortalityRate")
            PlantedArea_GAP_PACE_isPlantingDistance = ( 1 if request.form.get("PlantedArea_GAP_PACE_isPlantingDistance") else 0 )
            PlantedArea_GAP_PACE_isDiggingHoles = ( 1 if request.form.get("PlantedArea_GAP_PACE_isDiggingHoles") else 0 )
            PlantedArea_GAP_PACE_isSeparateTopSoil = ( 1 if request.form.get("PlantedArea_GAP_PACE_isSeparateTopSoil") else 0 )
            PlantedArea_GAP_PACE_isRemovingPlasticBag = ( 1 if request.form.get("PlantedArea_GAP_PACE_isRemovingPlasticBag") else 0 )
            PlantedArea_GAP_PACE_isApplyingMulching = ( 1 if request.form.get("PlantedArea_GAP_PACE_isApplyingMulching") else 0 )
            PlantedArea_GAP_PACE_isShadingEstablishment = ( 1 if request.form.get("PlantedArea_GAP_PACE_isShadingEstablishment") else 0 )
            
               
            PlantedArea_SOI_isOwn = ( 1 if request.form.get("PlantedArea_SOI")=="own" else 0 )
            PlantedArea_SOI_isGovAgency = ( 1 if request.form.get("PlantedArea_SOI")=="gov" else 0 )
            PlantedArea_SOI_isLoan = ( 1 if request.form.get("PlantedArea_SOI")=="loan" else 0 )
            PlantedArea_SOI_Others = ( request.form.get('PlantedArea_SOI_others') if request.form.get("PlantedArea_SOI")=="others" else 0 )
            
            PlantedArea_GAP_SANM_isBasal = ( 1 if request.form.get("PlantedArea_GAP_SANM_isBasal") else 0 )
            PlantedArea_GAP_SANM_isFoliar = ( 1 if request.form.get("PlantedArea_GAP_SANM_isFoliar") else 0 )
            PlantedArea_GAP_SANM_isPesticide = ( 1 if request.form.get("PlantedArea_GAP_SANM_isPesticide") else 0 )
            PlantedArea_GAP_SANM_isPlantBased = ( 1 if request.form.get("PlantedArea_GAP_SANM_isPlantBased") else 0 )
            
            PlantedArea_ValidatorName = request.form.get("PlantedArea_ValidatorName")
            PlantedArea_DateValidation = request.form.get("PlantedArea_DateValidation")
            PlantedArea_ValidationRemarks = request.form.get("PlantedArea_ValidationRemarks")
            
            AcknowOfFarmIntensification_NotPartOfMG = ( 1 if request.form.get("AcknowOfFarmIntensification_NotPartOfMG") else 0 )
            AcknowOfFarmIntensification_TargetArea = request.form.get("AcknowOfFarmIntensification_TargetArea")
            AcknowOfFarmIntensification_Name = request.form.get("AcknowOfFarmIntensification_Name")
            AcknowOfFarmIntensification_Qty = request.form.get("AcknowOfFarmIntensification_Qty")
            AcknowOfFarmIntensification_DateReceived = request.form.get("AcknowOfFarmIntensification_DateReceived")
            AcknowOfFarmIntensification_ValidatorName = request.form.get("AcknowOfFarmIntensification_ValidatorName")
            AcknowOfFarmIntensification_DateValidation = request.form.get("AcknowOfFarmIntensification_DateValidation")
            AcknowOfFarmIntensification_ValidationRemarks = request.form.get("AcknowOfFarmIntensification_ValidationRemarks")
            PlantedAreaForIntens_NotPartOfMG = ( 1 if request.form.get("PlantedAreaForIntens_NotPartOfMG") else 0 )
            PlantedAreaForIntens_PlantingTimeline = request.form.get("PlantedAreaForIntens_PlantingTimeline")
            PlantedAreaForIntens_TotalAreaPlanted = request.form.get("PlantedAreaForIntens_TotalAreaPlanted")
            PlantedAreaForIntens_NoOfSeedlingsPlanted = request.form.get("PlantedAreaForIntens_NoOfSeedlingsPlanted")
            PlantedAreaForIntens_NoOfSeedlingsNotPlanted = request.form.get("PlantedAreaForIntens_NoOfSeedlingsNotPlanted")
            PlantedAreaForIntens_PlantingDistance = request.form.get("PlantedAreaForIntens_PlantingDistance")
            PlantedAreaForIntens_HoleDepth = request.form.get("PlantedAreaForIntens_HoleDepth")
            PlantedAreaForIntens_MortalityRate = request.form.get("PlantedAreaForIntens_MortalityRate")
            PlantedAreaForIntens_ValidatedBy = request.form.get("PlantedAreaForIntens_ValidatedBy")
            PlantedAreaForIntens_DateValidation = request.form.get("PlantedAreaForIntens_DateValidation")
            PlantedAreaForIntens_ValidationRemarks = request.form.get("PlantedAreaForIntens_ValidationRemarks")
            AcknowOfFarmRehab_NotPartOfMG = ( 1 if request.form.get("AcknowOfFarmRehab_NotPartOfMG") else 0 )
            AcknowOfFarmRehab_TargetArea = request.form.get("AcknowOfFarmRehab_TargetArea")
            AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear")
            AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw")
            AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife")
            AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape")
            AcknowOfFarmRehab_QtyRehabToolsReceived_Others = request.form.get("AcknowOfFarmRehab_QtyRehabToolsReceived_Others")
            AcknowOfFarmRehab_DateReceived = request.form.get("AcknowOfFarmRehab_DateReceived")
            AcknowOfFarmRehab_ValidatorName = request.form.get("AcknowOfFarmRehab_ValidatorName")
            AcknowOfFarmRehab_DateValidation = request.form.get("AcknowOfFarmRehab_DateValidation")
            AcknowOfFarmRehab_ValidationRemarks = request.form.get("AcknowOfFarmRehab_ValidationRemarks")
            RehabFarmArea_GAP_NotParOfMG = ( 1 if request.form.get("RehabFarmArea_GAP_NotParOfMG") else 0 )
            RehabFarmArea_GAP_RehabTimeline = request.form.get("RehabFarmArea_GAP_RehabTimeline")
            RehabFarmArea_GAP_TotalAreaRehab = request.form.get("RehabFarmArea_GAP_TotalAreaRehab")
            RehabFarmArea_GAP_NoOfTreesRehab = request.form.get("RehabFarmArea_GAP_NoOfTreesRehab")
            RehabFarmArea_GAP_isShadingMaintenance = request.form.get("RehabFarmArea_GAP_isShadingMaintenance")
            RehabFarmArea_GAP_isTipPruning = request.form.get("RehabFarmArea_GAP_isTipPruning")
            RehabFarmArea_GAP_isShapePruning = request.form.get("RehabFarmArea_GAP_isShapePruning")
            RehabFarmArea_GAP_isAccessPruning = request.form.get("RehabFarmArea_GAP_isAccessPruning")
            RehabFarmArea_GAP_isSanitaryPruning = request.form.get("RehabFarmArea_GAP_isSanitaryPruning")
            RehabFarmArea_GAP_isMaintenancePruning = request.form.get("RehabFarmArea_GAP_isMaintenancePruning")
            RehabFarmArea_GAP_isUsingApprTools = request.form.get("RehabFarmArea_GAP_isUsingApprTools")
            RehabFarmArea_GAP_isFarmRecordKeeping = request.form.get("RehabFarmArea_GAP_isFarmRecordKeeping")
            RehabFarmArea_GAP_isChupon = request.form.get("RehabFarmArea_GAP_isChupon")
            RehabFarmArea_GAP_isFoliarFertilizer = request.form.get("RehabFarmArea_GAP_isFoliarFertilizer")
            RehabFarmArea_GAP_isSealing = request.form.get("RehabFarmArea_GAP_isSealing")
            RehabFarmArea_GAP_isObservance = request.form.get("RehabFarmArea_GAP_isObservance")
            RehabFarmArea_GAP_isGrafting = request.form.get("RehabFarmArea_GAP_isGrafting")
            RehabFarmArea_GAP_ValidatorName = request.form.get("RehabFarmArea_GAP_ValidatorName")
            RehabFarmArea_GAP_DateValidation = request.form.get("RehabFarmArea_GAP_DateValidation")
            RehabFarmArea_GAP_ValidationRemarks = request.form.get("RehabFarmArea_GAP_ValidationRemarks")
            AccessToProdInvestment_NotPartOfMG = ( 1 if request.form.get("AccessToProdInvestment_NotPartOfMG") else 0 )
            AccessToProdInvestment_TypeOfProdInvestments = request.form.get("AccessToProdInvestment_TypeOfProdInvestments")
            AccessToProdInvestment_ValidatorName = request.form.get("AccessToProdInvestment_ValidatorName")
            AccessToProdInvestment_DateValidation = request.form.get("AccessToProdInvestment_DateValidation")
            AccessToProdInvestment_ValidationRemarks = request.form.get("AccessToProdInvestment_ValidationRemarks")
            filename = ''

            
            updateMGIFO = ("UPDATE mg_implementation_shf SET ImpUnits_RCU = '{}', ImpUnits_PCU = '{}', DIP_Name = '{}', DIP_Commodity = '{}', MembershipToCOOP = '{}', FarmerInfo_Fname = '{}', FarmerInfo_Mname = '{}', FarmerInfo_Lname = '{}', FarmerInfo_ExtName = '{}', FarmerInfo_DOB = '{}', FarmerInfo_Sex = '{}', FarmerInfo_Sector_isPWD = '{}', FarmerInfo_Sector_isYouth = '{}', FarmerInfo_Sector_isIP = '{}', FarmerInfo_Sector_isSC = '{}', FarmerInfo_address_Region = '{}', FarmerInfo_address_Province = '{}', FarmerInfo_address_municipality = '{}', FarmerInfo_address_brgy = '{}', FarmerInfo_address_street = '{}', FarmerInfo_farmLoc_long = '{}', FarmerInfo_farmLoc_lat = '{}', Acknowledgement_NotPartOfMG = '{}', Acknowledgement_TargetArea = '{}', Acknowledgement_Name = '{}', Acknowledgement_Qty = '{}', Acknowledgement_DateReceived = '{}', Acknowledgement_ValidatorName = '{}', Acknowledgement_DateValidation = '{}', Acknowledgement_Remarks = '{}', PlantedArea_GAP_NotPartOfMG = '{}', PlantedArea_GAP_PlantingTimeline = '{}', PlantedArea_GAP_TotalAreaPlanted = '{}', PlantedArea_GAP_NoOfSeedlingsPlanted = '{}', PlantedArea_GAP_NoOfSeedlingsNotPlanted = '{}', PlantedArea_GAP_MortalityRate = '{}', PlantedArea_GAP_PACE_isPlantingDistance = '{}', PlantedArea_GAP_PACE_isDiggingHoles = '{}', PlantedArea_GAP_PACE_isSeparateTopSoil = '{}', PlantedArea_GAP_PACE_isRemovingPlasticBag = '{}', PlantedArea_GAP_PACE_isApplyingMulching = '{}', PlantedArea_GAP_PACE_isShadingEstablishment = '{}', PlantedArea_SOI_isOwn = '{}', PlantedArea_SOI_isGovAgency = '{}', PlantedArea_SOI_isLoan = '{}', PlantedArea_SOI_Others = '{}', PlantedArea_GAP_SANM_isBasal = '{}', PlantedArea_GAP_SANM_isFoliar = '{}', PlantedArea_GAP_SANM_isPesticide = '{}', PlantedArea_GAP_SANM_isPlantBased = '{}', PlantedArea_ValidatorName = '{}', PlantedArea_DateValidation = '{}', PlantedArea_ValidationRemarks = '{}', AcknowOfFarmIntensification_NotPartOfMG = '{}', AcknowOfFarmIntensification_TargetArea = '{}', AcknowOfFarmIntensification_Name = '{}', AcknowOfFarmIntensification_Qty = '{}', AcknowOfFarmIntensification_DateReceived = '{}', AcknowOfFarmIntensification_ValidatorName = '{}', AcknowOfFarmIntensification_DateValidation = '{}', AcknowOfFarmIntensification_ValidationRemarks = '{}', PlantedAreaForIntens_NotPartOfMG = '{}', PlantedAreaForIntens_PlantingTimeline = '{}', PlantedAreaForIntens_TotalAreaPlanted = '{}', PlantedAreaForIntens_NoOfSeedlingsPlanted = '{}', PlantedAreaForIntens_NoOfSeedlingsNotPlanted = '{}', PlantedAreaForIntens_PlantingDistance = '{}', PlantedAreaForIntens_HoleDepth = '{}', PlantedAreaForIntens_MortalityRate = '{}', PlantedAreaForIntens_ValidatedBy = '{}', PlantedAreaForIntens_DateValidation = '{}', PlantedAreaForIntens_ValidationRemarks = '{}', AcknowOfFarmRehab_NotPartOfMG = '{}', AcknowOfFarmRehab_TargetArea = '{}', AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear = '{}', AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw = '{}', AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife = '{}', AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape = '{}', AcknowOfFarmRehab_QtyRehabToolsReceived_Others = '{}', AcknowOfFarmRehab_DateReceived = '{}', AcknowOfFarmRehab_ValidatorName = '{}', AcknowOfFarmRehab_DateValidation = '{}', AcknowOfFarmRehab_ValidationRemarks = '{}', RehabFarmArea_GAP_NotParOfMG = '{}', RehabFarmArea_GAP_RehabTimeline = '{}', RehabFarmArea_GAP_TotalAreaRehab = '{}', RehabFarmArea_GAP_NoOfTreesRehab = '{}', RehabFarmArea_GAP_isShadingMaintenance = '{}', RehabFarmArea_GAP_isTipPruning = '{}', RehabFarmArea_GAP_isShapePruning = '{}', RehabFarmArea_GAP_isAccessPruning = '{}', RehabFarmArea_GAP_isSanitaryPruning = '{}', RehabFarmArea_GAP_isMaintenancePruning = '{}', RehabFarmArea_GAP_isUsingApprTools = '{}', RehabFarmArea_GAP_isFarmRecordKeeping = '{}', RehabFarmArea_GAP_isChupon = '{}', RehabFarmArea_GAP_isFoliarFertilizer = '{}', RehabFarmArea_GAP_isSealing = '{}', RehabFarmArea_GAP_isObservance = '{}', RehabFarmArea_GAP_isGrafting = '{}', RehabFarmArea_GAP_ValidatorName = '{}', RehabFarmArea_GAP_DateValidation = '{}', RehabFarmArea_GAP_ValidationRemarks = '{}', AccessToProdInvestment_NotPartOfMG = '{}', AccessToProdInvestment_TypeOfProdInvestments = '{}', AccessToProdInvestment_ValidatorName = '{}', AccessToProdInvestment_DateValidation = '{}', AccessToProdInvestment_ValidationRemarks = '{}', DateModified = NOW() WHERE MGI_ID = {}".
                           format(ImpUnits_RCU , ImpUnits_PCU , DIP_Name , DIP_Commodity , MembershipToCOOP , FarmerInfo_Fname , FarmerInfo_Mname , FarmerInfo_Lname , FarmerInfo_ExtName , FarmerInfo_DOB , FarmerInfo_Sex , FarmerInfo_Sector_isPWD , FarmerInfo_Sector_isYouth , FarmerInfo_Sector_isIP , FarmerInfo_Sector_isSC , FarmerInfo_address_Region , FarmerInfo_address_Province , FarmerInfo_address_municipality , FarmerInfo_address_brgy , FarmerInfo_address_street , FarmerInfo_farmLoc_long , FarmerInfo_farmLoc_lat , Acknowledgement_NotPartOfMG , Acknowledgement_TargetArea , Acknowledgement_Name , Acknowledgement_Qty , Acknowledgement_DateReceived , Acknowledgement_ValidatorName , Acknowledgement_DateValidation , Acknowledgement_Remarks , PlantedArea_GAP_NotPartOfMG , PlantedArea_GAP_PlantingTimeline , PlantedArea_GAP_TotalAreaPlanted , PlantedArea_GAP_NoOfSeedlingsPlanted , PlantedArea_GAP_NoOfSeedlingsNotPlanted , PlantedArea_GAP_MortalityRate , PlantedArea_GAP_PACE_isPlantingDistance , PlantedArea_GAP_PACE_isDiggingHoles , PlantedArea_GAP_PACE_isSeparateTopSoil , PlantedArea_GAP_PACE_isRemovingPlasticBag , PlantedArea_GAP_PACE_isApplyingMulching , PlantedArea_GAP_PACE_isShadingEstablishment , PlantedArea_SOI_isOwn , PlantedArea_SOI_isGovAgency , PlantedArea_SOI_isLoan , PlantedArea_SOI_Others , PlantedArea_GAP_SANM_isBasal , PlantedArea_GAP_SANM_isFoliar , PlantedArea_GAP_SANM_isPesticide , PlantedArea_GAP_SANM_isPlantBased , PlantedArea_ValidatorName , PlantedArea_DateValidation , PlantedArea_ValidationRemarks , AcknowOfFarmIntensification_NotPartOfMG , AcknowOfFarmIntensification_TargetArea , AcknowOfFarmIntensification_Name , AcknowOfFarmIntensification_Qty , AcknowOfFarmIntensification_DateReceived , AcknowOfFarmIntensification_ValidatorName , AcknowOfFarmIntensification_DateValidation , AcknowOfFarmIntensification_ValidationRemarks , PlantedAreaForIntens_NotPartOfMG , PlantedAreaForIntens_PlantingTimeline , PlantedAreaForIntens_TotalAreaPlanted , PlantedAreaForIntens_NoOfSeedlingsPlanted , PlantedAreaForIntens_NoOfSeedlingsNotPlanted , PlantedAreaForIntens_PlantingDistance , PlantedAreaForIntens_HoleDepth , PlantedAreaForIntens_MortalityRate , PlantedAreaForIntens_ValidatedBy , PlantedAreaForIntens_DateValidation , PlantedAreaForIntens_ValidationRemarks , AcknowOfFarmRehab_NotPartOfMG , AcknowOfFarmRehab_TargetArea , AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear , AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw , AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife , AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape , AcknowOfFarmRehab_QtyRehabToolsReceived_Others , AcknowOfFarmRehab_DateReceived , AcknowOfFarmRehab_ValidatorName , AcknowOfFarmRehab_DateValidation , AcknowOfFarmRehab_ValidationRemarks , RehabFarmArea_GAP_NotParOfMG , RehabFarmArea_GAP_RehabTimeline , RehabFarmArea_GAP_TotalAreaRehab , RehabFarmArea_GAP_NoOfTreesRehab , RehabFarmArea_GAP_isShadingMaintenance , RehabFarmArea_GAP_isTipPruning , RehabFarmArea_GAP_isShapePruning , RehabFarmArea_GAP_isAccessPruning , RehabFarmArea_GAP_isSanitaryPruning , RehabFarmArea_GAP_isMaintenancePruning , RehabFarmArea_GAP_isUsingApprTools , RehabFarmArea_GAP_isFarmRecordKeeping , RehabFarmArea_GAP_isChupon , RehabFarmArea_GAP_isFoliarFertilizer , RehabFarmArea_GAP_isSealing , RehabFarmArea_GAP_isObservance , RehabFarmArea_GAP_isGrafting , RehabFarmArea_GAP_ValidatorName , RehabFarmArea_GAP_DateValidation , RehabFarmArea_GAP_ValidationRemarks , AccessToProdInvestment_NotPartOfMG , AccessToProdInvestment_TypeOfProdInvestments , AccessToProdInvestment_ValidatorName , AccessToProdInvestment_DateValidation , AccessToProdInvestment_ValidationRemarks , id))
            
            update = db.do(updateMGIFO)
            # print("ImpUnits_RCU:", ImpUnits_RCU, type(ImpUnits_RCU))
            # print("ImpUnits_PCU:", ImpUnits_PCU, type(ImpUnits_PCU))
            # print("id:", id, type(id))
            # print("SQL : ", updateMGIFO)
            
            if (update["response"]=="error"):
                flash(f"An error occured!", "error")
                # print(str(update))
            else:
                flash(f"Successfully saved the edited data!", "success")
            return redirect("/mis-v4/core-mg-implementation")
        
        
    @app.route('/mgimplementation_form_data', methods=['GET'])
    def mgimplementation_form_data():
        
        if not is_on_session():
            return redirect("/login")
        
        action = request.args.get('action')
        form = request.args.get('form')
        id = request.args.get('id')
        data = {}
        
        # print("ID Here...... : ",id," action here...... : ",action)
        
        if (action=="edit" or action=="view"):
            if (form=="fo"):
                if (id):
                    row = rapid_sql.select("SELECT * FROM mg_implementation_fo WHERE MGI_ID = {}".format(id))
                    data['data'] = row
                else:
                    flash(f"No Record Identifier Found!", "error")
                    return redirect("/mis-v4/core-mg-implementation")
            elif (form=="shf"):
                if (id):
                    row = rapid_sql.select("SELECT * FROM mg_implementation_shf WHERE MGI_ID = {}".format(id))
                    data['data'] = row
                else:
                    flash(f"No Record Identifier Found!", "error")
                    return redirect("/mis-v4/core-mg-implementation")
                
        dip = rapid_sql.select("SELECT * FROM dcf_prep_review_aprv_status WHERE form_1_name_dip <> '' ORDER BY form_1_name_dip ASC")
        pfb = rapid_sql.select("SELECT * FROM form_b WHERE organization_registered_name <> '' ORDER BY organization_registered_name ASC")
        
        data['dip'] = dip
        data['pfb'] = pfb
        data['form'] = form
        
        # print(f"data here...... : ", data)
        return jsonify(data)
    
    @app.route('/mgimplementation_data', methods=['GET'])
    def mgimplementation_data():
        
        if(session["USER_DATA"][0]['security_group']==0): return redirect("/warning?type=user-no-role");
  
        type = request.args.get("type") 
        data = []
        if ( type == "fo" ):
            data = rapid_sql.select("SELECT *, FOInfo_FOName AS Name,'fo' AS 'MGI_Type' FROM mg_implementation_fo WHERE isDeleted=0")
        else:
            data = rapid_sql.select("SELECT *, CONCAT(FarmerInfo_Fname, ' ', FarmerInfo_Mname, ' ', FarmerInfo_Lname,' ', FarmerInfo_ExtName) AS Name,'shf' AS 'MGI_Type' FROM mg_implementation_shf WHERE isDeleted=0")
        
        result = {
            'data' : data,
            'user_detail' : session["USER_DATA"][0]
        }

        return jsonify(result)

    @app.route('/export_mgimplementation', methods=['GET','POST'])
    def export_mgimplementation():
        mgi_type = request.args.get("type")
        if mgi_type == "fo":
            data = rapid_sql.select("SELECT * FROM mg_implementation_fo")
            df = pd.json_normalize(data).astype(str)

            new_column_names = 'MGI_ID,UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,FOInfo_FOName,FOInfo_Address_Region,FOInfo_Address_Province,FOInfo_Address_Municipality,FOInfo_Address_Brgy,FOInfo_Address_Street,FOInfo_Head_Fname,FOInfo_Head_Mname,FOInfo_Head_Lname,FOInfo_Head_Extname,FOInfo_Head_Sex,FOInfo_Head_Sector,AcknowOfFarmExp_TargetArea,AcknowOfFarmExp_ItemName,AcknowOfFarmExp_QtyReceived,AcknowOfFarmExp_DateReceived,AcknowOfFarmExp_FarmerRep_Fname,AcknowOfFarmExp_FarmerRep_Mname,AcknowOfFarmExp_FarmerRep_Lname,AcknowOfFarmExp_FarmerRep_Extname,AcknowOfFarmExp_FarmerRep_Designation,AcknowOfFarmExp_Witness1_ValidatorName,AcknowOfFarmExp_Witness1_ValidationDate,AcknowOfFarmExp_Witness1_Designation,AcknowOfFarmExp_Witness1_DateWitnessed,AcknowOfFarmExp_Witness2_ValidatorName,AcknowOfFarmExp_Witness2_ValidationDate,AcknowOfFarmExp_Witness2_Remarks,AcknowOfFarmExp_Witness1Raw_ValidatorName,AcknowOfFarmExp_Witness1Raw_ValidationDate,AcknowOfFarmExp_Witness1Raw_Designation,AcknowOfFarmExp_Witness1Raw_DateWitnessed,AcknowOfFarmExp_Witness2Raw_ValidatorName,AcknowOfFarmExp_Witness2Raw_ValidationDate,AcknowOfFarmExp_Witness2Raw_Remarks,AcknowOfFarmInten_TargetArea,AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived,AcknowOfFarmInten_DateReceived,AcknowOfFarmInten_FarmerRep_Fname,AcknowOfFarmInten_FarmerRep_Mname,AcknowOfFarmInten_FarmerRep_Lname,AcknowOfFarmInten_FarmerRep_Extname,AcknowOfFarmInten_FarmerRep_Designation,AcknowOfFarmInten_Withness1_ValidatorName,AcknowOfFarmInten_Withness1_ValidationDate,AcknowOfFarmInten_Withness1_Designation,AcknowOfFarmInten_Withness1_DateWitnessed,AcknowOfFarmInten_Withness2_ValidatorName,AcknowOfFarmInten_Withness2_ValidationDate,AcknowOfFarmInten_Withness2_Remarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_FarmerRep_Fname,AcknowOfFarmRehab_FarmerRep_Mname,AcknowOfFarmRehab_FarmerRep_Lname,AcknowOfFarmRehab_FarmerRep_Extname,AcknowOfFarmRehab_FarmerRep_Designation,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_ValidationDate,AcknowOfFarmRehab_Remarks,AcknowOfProd_TypeOfProdInv,AcknowOfProd_QtyItemReceived,AcknowOfProd_DateReceived,AcknowOfProd_FarmerRep_Fname,AcknowOfProd_FarmerRep_Mname,AcknowOfProd_FarmerRep_Lname,AcknowOfProd_FarmerRep_Extname,AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,DateModified,filename' 

            # Adjust column names to match DataFrame shape
            new_column_names_list = new_column_names.split(',')
            if len(new_column_names_list) == len(df.columns):
                df.columns = new_column_names_list
            else:
                # Example: trim or pad to match
                df = df.iloc[:, :len(new_column_names_list)]
                df.columns = new_column_names_list

            file_path = c.RECORDS+'/objects/_temp_/mgi_fo_exported_file.xlsx'
            with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
                df.to_excel(writer, sheet_name='mgi_fo_exported_file', index=False)

                workbook = writer.book
                worksheet = writer.sheets['mgi_fo_exported_file']
                header_format = workbook.add_format({
                    'bold': True, 'text_wrap': True, 'valign': 'top',
                    'fg_color': '#00ace6', 'border': 1
                })
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                    column_width = max(df[value].astype(str).apply(len).max(), len(value)) + 1
                    worksheet.set_column(col_num, col_num, column_width)
            return send_file(file_path)
        elif mgi_type == "shf":
            data = rapid_sql.select("SELECT * FROM mg_implementation_shf")
            df = pd.json_normalize(data).astype(str)

            new_column_names = 'MGI_ID,UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,MembershipToCOOP,FarmerInfo_Fname,FarmerInfo_Mname,FarmerInfo_Lname,FarmerInfo_ExtName,FarmerInfo_DOB,FarmerInfo_Sex,FarmerInfo_isPWD,FarmerInfo_isYouth,FarmerInfo_isIP,FarmerInfo_isSC,FarmerInfo_address_Region,FarmerInfo_address_Province,FarmerInfo_address_municipality,FarmerInfo_address_brgy,FarmerInfo_address_street,FarmerInfo_farmLoc_long,FarmerInfo_farmLoc_lat,Acknowledgement_TargetArea,Acknowledgement_Name,Acknowledgement_Qty,Acknowledgement_DateReceived,Acknowledgement_ValidatorName,Acknowledgement_DateValidation,Acknowledgement_Remarks,PlantedArea_GAP_PACE_isPlantingDistance,PlantedArea_GAP_PACE_isDiggingHoles,PlantedArea_GAP_PACE_isSeparateTopSoil,PlantedArea_GAP_PACE_isRemovingPlasticBag,PlantedArea_GAP_PACE_isApplyingMulching,PlantedArea_GAP_PACE_isShadingEstablishment,PlantedArea_SOI_isOwn,PlantedArea_SOI_isGovAgency,PlantedArea_SOI_isLoan,PlantedArea_SOI_Others,PlantedArea_GAP_SANM_isBasal,PlantedArea_GAP_SANM_isFoliar,PlantedArea_GAP_SANM_isPesticide,PlantedArea_GAP_SANM_isPlantBased,PlantedArea_ValidatorName,PlantedArea_DateValidation,PlantedArea_ValidationRemarks,AcknowOfFarmIntensification_TargetArea,AcknowOfFarmIntensification_Name,AcknowOfFarmIntensification_Qty,AcknowOfFarmIntensification_DateReceived,AcknowOfFarmIntensification_ValidatorName,AcknowOfFarmIntensification_DateValidation,AcknowOfFarmIntensification_ValidationRemarks,PlantedAreaForIntens_PlantingTimeline,PlantedAreaForIntens_TotalAreaPlanted,PlantedAreaForIntens_NoOfSeedlingsPlanted,PlantedAreaForIntens_NoOfSeedlingsNotPlanted,PlantedAreaForIntens_PlantingDistance,PlantedAreaForIntens_HoleDepth,PlantedAreaForIntens_MortalityRate,PlantedAreaForIntens_ValidatedBy,PlantedAreaForIntens_DateValidation,PlantedAreaForIntens_ValidationRemarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw,AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife,AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape,AcknowOfFarmRehab_QtyRehabToolsReceived_Others,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_DateValidation,AcknowOfFarmRehab_ValidationRemarks,RehabFarmArea_GAP_isShadingMaintenance,RehabFarmArea_GAP_isTipPruning,RehabFarmArea_GAP_isShapePruning,RehabFarmArea_GAP_isAccessPruning,RehabFarmArea_GAP_isSanitaryPruning,RehabFarmArea_GAP_isMaintenancePruning,RehabFarmArea_GAP_isUsingApprTools,RehabFarmArea_GAP_isFarmRecordKeeping,RehabFarmArea_GAP_isChupon,RehabFarmArea_GAP_isFoliarFertilizer,RehabFarmArea_GAP_isSealing,RehabFarmArea_GAP_isObservance,RehabFarmArea_GAP_isGrafting,RehabFarmArea_GAP_ValidatorName,RehabFarmArea_GAP_DateValidation,RehabFarmArea_GAP_ValidationRemarks,AccessToProdInvestment_TypeOfProdInvestments,AccessToProdInvestment_ValidatorName,AccessToProdInvestment_DateValidation,AccessToProdInvestment_ValidationRemarks,DateModified,filename' 

            # Adjust column names to match DataFrame shape
            new_column_names_list = new_column_names.split(',')
            if len(new_column_names_list) == len(df.columns):
                df.columns = new_column_names_list
            else:
                # Example: trim or pad to match
                df = df.iloc[:, :len(new_column_names_list)]
                df.columns = new_column_names_list

            file_path = c.RECORDS+'/objects/_temp_/mgi_shf_exported_file.xlsx'
            with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
                df.to_excel(writer, sheet_name='mgi_shf_exported_file', index=False)

                workbook = writer.book
                worksheet = writer.sheets['mgi_shf_exported_file']
                header_format = workbook.add_format({
                    'bold': True, 'text_wrap': True, 'valign': 'top',
                    'fg_color': '#00ace6', 'border': 1
                })
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                    column_width = max(df[value].astype(str).apply(len).max(), len(value)) + 1
                    worksheet.set_column(col_num, col_num, column_width)
            return send_file(file_path)

    @app.route('/importcsv_mgimplementation',methods = ['GET','POST'])
    def importcsv_mgimplementation():
        _main.importcsv_mgimplementation(request)
        return redirect("/mis-v4/core-mg-implementation")
    
    def importcsv_mgimplementation(request):
        from datetime import datetime
        today = str(datetime.today()).replace("-", "").replace(" ", "").replace(":", "").replace(".", "")
        uploader = session["USER_DATA"][0]["id"]
        impType = request.form.get("import_type")   
        
        if request.method == "POST":
            try:
                files = request.files
                for file in files:
                    f = files[file]
                    global UPLOAD_NAME
                    UPLOAD_NAME = str(uploader) + "#" + str(today) + "#" + secure_filename(f.filename)
                    f.save(os.path.join(c.RECORDS + "/objects/spreadsheets_mgi/queued/", UPLOAD_NAME))
                    _main.excel_upload_open_mgi(os.path.join(c.RECORDS + "/objects/spreadsheets_mgi/queued/", UPLOAD_NAME), impType)
            except IndexError:
                # flash("I am here!", "error")
                flash("Invalid file template!", "error")

        return redirect("/dcfspreadsheet")

    def safe_get(row, idx, default=None):
        return row[idx] if idx < len(row) and row[idx] is not None else default

    def excel_upload_open_mgi(path, impType):  
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        insert = None
        # book = xlrd.open_workbook(path)
        # sheet = book.sheet_by_index(0)
        # data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]

        # data = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=True)]
        # data = [list(row) for row in sheet.iter_rows(values_only=True)]

        # header = data[4]
    
        if impType == "shf":
            if "DDN_SHFs" in wb.sheetnames:
                    sheet = wb["DDN_SHFs"]
            else:
                flash("Invalid file template!", "error")
                return "done:S"

            data = [list(row) for row in sheet.iter_rows(values_only=True)]
            
            for row in data[9:]:
                print("Row length:", len(row), "Row data:", row)
                print(" Row[2] ==>", row[2])
                print(" Row[103] ==>", row[103])
                if not any(cell and str(cell).strip() for cell in row):
                    continue

                UploadedBy = session["USER_DATA"][0]['id']
                
                ImpUnits_RCU = _main.safe_get(row,2,"")
                ImpUnits_PCU = _main.safe_get(row,3,"")
                DIP_Name = _main.safe_get(row,4,"")
                DIP_Commodity = _main.safe_get(row,5,"")
                MembershipToCOOP = _main.safe_get(row,6,"")
                FarmerInfo_Fname = _main.safe_get(row,7,"")
                FarmerInfo_Mname = _main.safe_get(row,8,"")
                FarmerInfo_Lname = _main.safe_get(row,9,"")
                FarmerInfo_ExtName = _main.safe_get(row,10,"")
                FarmerInfo_DOB = _main.safe_get(row,11,"")
                
                FarmerInfo_Sex = ''
                if ( _main.safe_get(row,12,"")==1 ):
                    FarmerInfo_Sex = "M"
                elif ( _main.safe_get(row,13,"")==1 ):
                    FarmerInfo_Sex = "F"
                
                FarmerInfo_Sector_isPWD = _main.safe_get(row,14,0)
                FarmerInfo_Sector_isYouth = _main.safe_get(row,15,0)
                FarmerInfo_Sector_isIP = _main.safe_get(row,16,0)
                FarmerInfo_Sector_isSC = _main.safe_get(row,17,0)
                FarmerInfo_address_Region = _main.safe_get(row,18,"")
                FarmerInfo_address_Province = _main.safe_get(row,19,"")
                FarmerInfo_address_municipality = _main.safe_get(row,20,"")
                FarmerInfo_address_brgy = _main.safe_get(row,21,"")
                FarmerInfo_address_street = _main.safe_get(row,22,"")
                FarmerInfo_farmLoc_long = _main.safe_get(row,23,"")
                FarmerInfo_farmLoc_lat = _main.safe_get(row,24,"") 
                Acknowledgement_TargetArea = _main.safe_get(row,25,"")
                Acknowledgement_Name = _main.safe_get(row,26,"")
                Acknowledgement_Qty = _main.safe_get(row,27,"")
                Acknowledgement_DateReceived = _main.safe_get(row,28,"")
                Acknowledgement_ValidatorName = _main.safe_get(row,29,"")
                Acknowledgement_DateValidation = _main.safe_get(row,30,"")
                Acknowledgement_Remarks = _main.safe_get(row,31,"")
                PlantedArea_GAP_PlantingTimeline = _main.safe_get(row,32,"")
                PlantedArea_GAP_TotalAreaPlanted = _main.safe_get(row,33,"")
                PlantedArea_GAP_NoOfSeedlingsPlanted = _main.safe_get(row,34,"")
                PlantedArea_GAP_NoOfSeedlingsNotPlanted = _main.safe_get(row,35,"")
                PlantedArea_GAP_MortalityRate = _main.safe_get(row,36,"")
                PlantedArea_GAP_PACE_isPlantingDistance = _main.safe_get(row,37,0)
                PlantedArea_GAP_PACE_isDiggingHoles = _main.safe_get(row,38,0)
                PlantedArea_GAP_PACE_isSeparateTopSoil = _main.safe_get(row,39,0)
                PlantedArea_GAP_PACE_isRemovingPlasticBag = _main.safe_get(row,40,0)
                PlantedArea_GAP_PACE_isApplyingMulching = _main.safe_get(row,41,0)
                PlantedArea_GAP_PACE_isShadingEstablishment = _main.safe_get(row,42,0)
                PlantedArea_SOI_isOwn = _main.safe_get(row,43,0)
                PlantedArea_SOI_isGovAgency = _main.safe_get(row,44,0)
                PlantedArea_SOI_isLoan = _main.safe_get(row,45,0)
                PlantedArea_SOI_Others = _main.safe_get(row,46,"")
                PlantedArea_GAP_SANM_isBasal = _main.safe_get(row,47,0)
                PlantedArea_GAP_SANM_isFoliar = _main.safe_get(row,48,0)
                PlantedArea_GAP_SANM_isPesticide = _main.safe_get(row,49,0)
                PlantedArea_GAP_SANM_isPlantBased = _main.safe_get(row,50,0)
                PlantedArea_ValidatorName = _main.safe_get(row,51,"")
                PlantedArea_DateValidation = _main.safe_get(row,52,"")
                PlantedArea_ValidationRemarks = _main.safe_get(row,53,"")
                AcknowOfFarmIntensification_TargetArea = _main.safe_get(row,54,"")
                AcknowOfFarmIntensification_Name = _main.safe_get(row,55,"")
                AcknowOfFarmIntensification_Qty = _main.safe_get(row,56,"")
                AcknowOfFarmIntensification_DateReceived = _main.safe_get(row,57,"")
                AcknowOfFarmIntensification_ValidatorName = _main.safe_get(row,58,"")
                AcknowOfFarmIntensification_DateValidation = _main.safe_get(row,59,"")
                AcknowOfFarmIntensification_ValidationRemarks = _main.safe_get(row,60,"")
                PlantedAreaForIntens_PlantingTimeline = _main.safe_get(row,61,"")
                PlantedAreaForIntens_TotalAreaPlanted = _main.safe_get(row,62,"")
                PlantedAreaForIntens_NoOfSeedlingsPlanted = _main.safe_get(row,63,"")
                PlantedAreaForIntens_NoOfSeedlingsNotPlanted = _main.safe_get(row,64,"")
                PlantedAreaForIntens_PlantingDistance = _main.safe_get(row,65,"")
                PlantedAreaForIntens_HoleDepth = _main.safe_get(row,66,"")
                PlantedAreaForIntens_MortalityRate = _main.safe_get(row,67,"")
                PlantedAreaForIntens_ValidatedBy = _main.safe_get(row,68,"")
                PlantedAreaForIntens_DateValidation = _main.safe_get(row,69,"")
                PlantedAreaForIntens_ValidationRemarks = _main.safe_get(row,70,"")
                AcknowOfFarmRehab_TargetArea = _main.safe_get(row,71,"")
                AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear = _main.safe_get(row,72,"")
                AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw = _main.safe_get(row,73,"")
                AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife = _main.safe_get(row,74,"")
                AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape = _main.safe_get(row,75,"")
                AcknowOfFarmRehab_QtyRehabToolsReceived_Others = _main.safe_get(row,76,"")
                AcknowOfFarmRehab_DateReceived = _main.safe_get(row,77,"")
                AcknowOfFarmRehab_ValidatorName = _main.safe_get(row,78,"")
                AcknowOfFarmRehab_DateValidation = _main.safe_get(row,79,"")
                AcknowOfFarmRehab_ValidationRemarks = _main.safe_get(row,80,"")
                RehabFarmArea_GAP_RehabTimeline = _main.safe_get(row,81,"")
                RehabFarmArea_GAP_TotalAreaRehab = _main.safe_get(row,82,"")
                RehabFarmArea_GAP_NoOfTreesRehab = _main.safe_get(row,83,"")
                RehabFarmArea_GAP_isShadingMaintenance = _main.safe_get(row,84,0)
                RehabFarmArea_GAP_isTipPruning = _main.safe_get(row,85,0)
                RehabFarmArea_GAP_isShapePruning = _main.safe_get(row,86,0)
                RehabFarmArea_GAP_isAccessPruning = _main.safe_get(row,87,0)
                RehabFarmArea_GAP_isSanitaryPruning = _main.safe_get(row,88,0)
                RehabFarmArea_GAP_isMaintenancePruning = _main.safe_get(row,89,0)
                RehabFarmArea_GAP_isUsingApprTools = _main.safe_get(row,90,0)
                RehabFarmArea_GAP_isFarmRecordKeeping = _main.safe_get(row,91,0)
                RehabFarmArea_GAP_isChupon = _main.safe_get(row,92,0)
                RehabFarmArea_GAP_isFoliarFertilizer = _main.safe_get(row,93,0)
                RehabFarmArea_GAP_isSealing = _main.safe_get(row,94,0)
                RehabFarmArea_GAP_isObservance = _main.safe_get(row,95,0)
                RehabFarmArea_GAP_isGrafting = _main.safe_get(row,96,0)
                RehabFarmArea_GAP_ValidatorName = _main.safe_get(row,97,"")
                RehabFarmArea_GAP_DateValidation = _main.safe_get(row,98,"")
                RehabFarmArea_GAP_ValidationRemarks = _main.safe_get(row,99,"")
                AccessToProdInvestment_TypeOfProdInvestments = _main.safe_get(row,100,"")
                AccessToProdInvestment_ValidatorName = _main.safe_get(row,101,"")
                AccessToProdInvestment_DateValidation = _main.safe_get(row,102,"")
                AccessToProdInvestment_ValidationRemarks = _main.safe_get(row,103,"")
                filename = UPLOAD_NAME
                
                # ImpUnits_RCU, ImpUnits_PCU, DIP_Name, DIP_Commodity, MembershipToCOOP, FarmerInfo_Fname, FarmerInfo_Mname, FarmerInfo_Lname, FarmerInfo_ExtName, FarmerInfo_DOB, FarmerInfo_Sex, FarmerInfo_isPWD, FarmerInfo_isYouth, FarmerInfo_isIP, FarmerInfo_isSC, FarmerInfo_address_Region, FarmerInfo_address_Province, FarmerInfo_address_municipality, FarmerInfo_address_brgy, FarmerInfo_address_street, FarmerInfo_farmLoc_long, FarmerInfo_farmLoc_lat, Acknowledgement_TargetArea, Acknowledgement_Name, Acknowledgement_Qty, Acknowledgement_DateReceived, Acknowledgement_ValidatorName, Acknowledgement_DateValidation, Acknowledgement_Remarks, PlantedArea_GAP_PlantingTimeline, PlantedArea_GAP_TotalAreaPlanted, PlantedArea_GAP_NoOfSeedlingsPlanted, PlantedArea_GAP_NoOfSeedlingsNotPlanted, PlantedArea_GAP_MortalityRate, PlantedArea_GAP_PACE_isPlantingDistance, PlantedArea_GAP_PACE_isDiggingHoles, PlantedArea_GAP_PACE_isSeparateTopSoil, PlantedArea_GAP_PACE_isRemovingPlasticBag, PlantedArea_GAP_PACE_isApplyingMulching, PlantedArea_GAP_PACE_isShadingEstablishment, PlantedArea_SOI_isOwn, PlantedArea_SOI_isGovAgency, PlantedArea_SOI_isLoan, PlantedArea_SOI_Others, PlantedArea_GAP_SANM_isBasal, PlantedArea_GAP_SANM_isFoliar, PlantedArea_GAP_SANM_isPesticide, PlantedArea_GAP_SANM_isPlantBased, PlantedArea_ValidatorName, PlantedArea_DateValidation, PlantedArea_ValidationRemarks, AcknowOfFarmIntensification_TargetArea, AcknowOfFarmIntensification_Name, AcknowOfFarmIntensification_Qty, AcknowOfFarmIntensification_DateReceived, AcknowOfFarmIntensification_ValidatorName, AcknowOfFarmIntensification_DateValidation, AcknowOfFarmIntensification_ValidationRemarks, PlantedAreaForIntens_PlantingTimeline, PlantedAreaForIntens_TotalAreaPlanted, PlantedAreaForIntens_NoOfSeedlingsPlanted, PlantedAreaForIntens_NoOfSeedlingsNotPlanted, PlantedAreaForIntens_PlantingDistance, PlantedAreaForIntens_HoleDepth, PlantedAreaForIntens_MortalityRate, PlantedAreaForIntens_ValidatedBy, PlantedAreaForIntens_DateValidation, PlantedAreaForIntens_ValidationRemarks, AcknowOfFarmRehab_TargetArea, AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear, AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw, AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife, AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape, AcknowOfFarmRehab_QtyRehabToolsReceived_Others, AcknowOfFarmRehab_DateReceived, AcknowOfFarmRehab_ValidatorName, AcknowOfFarmRehab_DateValidation, AcknowOfFarmRehab_ValidationRemarks, RehabFarmArea_GAP_RehabTimeline, RehabFarmArea_GAP_TotalAreaRehab, RehabFarmArea_GAP_NoOfTreesRehab, RehabFarmArea_GAP_isShadingMaintenance, RehabFarmArea_GAP_isTipPruning, RehabFarmArea_GAP_isShapePruning, RehabFarmArea_GAP_isAccessPruning, RehabFarmArea_GAP_isSanitaryPruning, RehabFarmArea_GAP_isMaintenancePruning, RehabFarmArea_GAP_isUsingApprTools, RehabFarmArea_GAP_isFarmRecordKeeping, RehabFarmArea_GAP_isChupon, RehabFarmArea_GAP_isFoliarFertilizer, RehabFarmArea_GAP_isSealing, RehabFarmArea_GAP_isObservance, RehabFarmArea_GAP_isGrafting, RehabFarmArea_GAP_ValidatorName, RehabFarmArea_GAP_DateValidation, RehabFarmArea_GAP_ValidationRemarks, AccessToProdInvestment_TypeOfProdInvestments, AccessToProdInvestment_ValidatorName, AccessToProdInvestment_DateValidation, AccessToProdInvestment_ValidationRemarks, 
                
                # querycsv = ("INSERT INTO mg_implementation_shf (UploadedBy,ImpUnits_RCU, ImpUnits_PCU, DIP_Name, DIP_Commodity, MembershipToCOOP, FarmerInfo_Fname, FarmerInfo_Mname, FarmerInfo_Lname, FarmerInfo_ExtName, FarmerInfo_DOB, FarmerInfo_Sex, FarmerInfo_Sector_isPWD, FarmerInfo_Sector_isYouth, FarmerInfo_Sector_isIP, FarmerInfo_Sector_isSC, FarmerInfo_address_Region, FarmerInfo_address_Province, FarmerInfo_address_municipality, FarmerInfo_address_brgy, FarmerInfo_address_street, FarmerInfo_farmLoc_long, FarmerInfo_farmLoc_lat) VALUES('{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}','{}', '{}' )").format(UploadedBy,ImpUnits_RCU, ImpUnits_PCU, DIP_Name, DIP_Commodity,MembershipToCOOP, FarmerInfo_Fname, FarmerInfo_Mname, FarmerInfo_Lname, FarmerInfo_ExtName, FarmerInfo_DOB, FarmerInfo_Sex, FarmerInfo_Sector_isPWD, FarmerInfo_Sector_isYouth, FarmerInfo_Sector_isIP, FarmerInfo_Sector_isSC, FarmerInfo_address_Region, FarmerInfo_address_Province, FarmerInfo_address_municipality, FarmerInfo_address_brgy, FarmerInfo_address_street, FarmerInfo_farmLoc_long, FarmerInfo_farmLoc_lat)
                querycsv = ("INSERT INTO mg_implementation_shf ( UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,MembershipToCOOP,FarmerInfo_Fname,FarmerInfo_Mname,FarmerInfo_Lname,FarmerInfo_ExtName,FarmerInfo_DOB,FarmerInfo_Sex,FarmerInfo_Sector_isPWD,FarmerInfo_Sector_isYouth,FarmerInfo_Sector_isIP,FarmerInfo_Sector_isSC,FarmerInfo_address_Region,FarmerInfo_address_Province,FarmerInfo_address_municipality,FarmerInfo_address_brgy,FarmerInfo_address_street,FarmerInfo_farmLoc_long,FarmerInfo_farmLoc_lat,Acknowledgement_TargetArea,Acknowledgement_Name,Acknowledgement_Qty,Acknowledgement_DateReceived,Acknowledgement_ValidatorName,Acknowledgement_DateValidation,Acknowledgement_Remarks,PlantedArea_GAP_PlantingTimeline,PlantedArea_GAP_TotalAreaPlanted,PlantedArea_GAP_NoOfSeedlingsPlanted,PlantedArea_GAP_NoOfSeedlingsNotPlanted,PlantedArea_GAP_MortalityRate,PlantedArea_GAP_PACE_isPlantingDistance,PlantedArea_GAP_PACE_isDiggingHoles,PlantedArea_GAP_PACE_isSeparateTopSoil,PlantedArea_GAP_PACE_isRemovingPlasticBag,PlantedArea_GAP_PACE_isApplyingMulching,PlantedArea_GAP_PACE_isShadingEstablishment,PlantedArea_SOI_isOwn,PlantedArea_SOI_isGovAgency,PlantedArea_SOI_isLoan,PlantedArea_SOI_Others,PlantedArea_GAP_SANM_isBasal,PlantedArea_GAP_SANM_isFoliar,PlantedArea_GAP_SANM_isPesticide,PlantedArea_GAP_SANM_isPlantBased,PlantedArea_ValidatorName,PlantedArea_DateValidation,PlantedArea_ValidationRemarks,AcknowOfFarmIntensification_TargetArea,AcknowOfFarmIntensification_Name,AcknowOfFarmIntensification_Qty,AcknowOfFarmIntensification_DateReceived,AcknowOfFarmIntensification_ValidatorName,AcknowOfFarmIntensification_DateValidation,AcknowOfFarmIntensification_ValidationRemarks,PlantedAreaForIntens_PlantingTimeline,PlantedAreaForIntens_TotalAreaPlanted,PlantedAreaForIntens_NoOfSeedlingsPlanted,PlantedAreaForIntens_NoOfSeedlingsNotPlanted,PlantedAreaForIntens_PlantingDistance,PlantedAreaForIntens_HoleDepth,PlantedAreaForIntens_MortalityRate,PlantedAreaForIntens_ValidatedBy,PlantedAreaForIntens_DateValidation,PlantedAreaForIntens_ValidationRemarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw,AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife,AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape,AcknowOfFarmRehab_QtyRehabToolsReceived_Others,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_DateValidation,AcknowOfFarmRehab_ValidationRemarks,RehabFarmArea_GAP_RehabTimeline,RehabFarmArea_GAP_TotalAreaRehab,RehabFarmArea_GAP_NoOfTreesRehab,RehabFarmArea_GAP_isShadingMaintenance,RehabFarmArea_GAP_isTipPruning,RehabFarmArea_GAP_isShapePruning,RehabFarmArea_GAP_isAccessPruning,RehabFarmArea_GAP_isSanitaryPruning,RehabFarmArea_GAP_isMaintenancePruning,RehabFarmArea_GAP_isUsingApprTools,RehabFarmArea_GAP_isFarmRecordKeeping,RehabFarmArea_GAP_isChupon,RehabFarmArea_GAP_isFoliarFertilizer,RehabFarmArea_GAP_isSealing,RehabFarmArea_GAP_isObservance,RehabFarmArea_GAP_isGrafting,RehabFarmArea_GAP_ValidatorName,RehabFarmArea_GAP_DateValidation,RehabFarmArea_GAP_ValidationRemarks,AccessToProdInvestment_TypeOfProdInvestments,AccessToProdInvestment_ValidatorName,AccessToProdInvestment_DateValidation,AccessToProdInvestment_ValidationRemarks,filename) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".
                format(UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,MembershipToCOOP,FarmerInfo_Fname,FarmerInfo_Mname,FarmerInfo_Lname,FarmerInfo_ExtName,FarmerInfo_DOB,FarmerInfo_Sex,FarmerInfo_Sector_isPWD,FarmerInfo_Sector_isYouth,FarmerInfo_Sector_isIP,FarmerInfo_Sector_isSC,FarmerInfo_address_Region,FarmerInfo_address_Province,FarmerInfo_address_municipality,FarmerInfo_address_brgy,FarmerInfo_address_street,FarmerInfo_farmLoc_long,FarmerInfo_farmLoc_lat,Acknowledgement_TargetArea,Acknowledgement_Name,Acknowledgement_Qty,Acknowledgement_DateReceived,Acknowledgement_ValidatorName,Acknowledgement_DateValidation,Acknowledgement_Remarks,PlantedArea_GAP_PlantingTimeline,PlantedArea_GAP_TotalAreaPlanted,PlantedArea_GAP_NoOfSeedlingsPlanted,PlantedArea_GAP_NoOfSeedlingsNotPlanted,PlantedArea_GAP_MortalityRate,PlantedArea_GAP_PACE_isPlantingDistance,PlantedArea_GAP_PACE_isDiggingHoles,PlantedArea_GAP_PACE_isSeparateTopSoil,PlantedArea_GAP_PACE_isRemovingPlasticBag,PlantedArea_GAP_PACE_isApplyingMulching,PlantedArea_GAP_PACE_isShadingEstablishment,PlantedArea_SOI_isOwn,PlantedArea_SOI_isGovAgency,PlantedArea_SOI_isLoan,PlantedArea_SOI_Others,PlantedArea_GAP_SANM_isBasal,PlantedArea_GAP_SANM_isFoliar,PlantedArea_GAP_SANM_isPesticide,PlantedArea_GAP_SANM_isPlantBased,PlantedArea_ValidatorName,PlantedArea_DateValidation,PlantedArea_ValidationRemarks,AcknowOfFarmIntensification_TargetArea,AcknowOfFarmIntensification_Name,AcknowOfFarmIntensification_Qty,AcknowOfFarmIntensification_DateReceived,AcknowOfFarmIntensification_ValidatorName,AcknowOfFarmIntensification_DateValidation,AcknowOfFarmIntensification_ValidationRemarks,PlantedAreaForIntens_PlantingTimeline,PlantedAreaForIntens_TotalAreaPlanted,PlantedAreaForIntens_NoOfSeedlingsPlanted,PlantedAreaForIntens_NoOfSeedlingsNotPlanted,PlantedAreaForIntens_PlantingDistance,PlantedAreaForIntens_HoleDepth,PlantedAreaForIntens_MortalityRate,PlantedAreaForIntens_ValidatedBy,PlantedAreaForIntens_DateValidation,PlantedAreaForIntens_ValidationRemarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningShear,AcknowOfFarmRehab_QtyRehabToolsReceived_PrunningSaw,AcknowOfFarmRehab_QtyRehabToolsReceived_BuddingKnife,AcknowOfFarmRehab_QtyRehabToolsReceived_GraftingTape,AcknowOfFarmRehab_QtyRehabToolsReceived_Others,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_DateValidation,AcknowOfFarmRehab_ValidationRemarks,RehabFarmArea_GAP_RehabTimeline,RehabFarmArea_GAP_TotalAreaRehab,RehabFarmArea_GAP_NoOfTreesRehab,RehabFarmArea_GAP_isShadingMaintenance,RehabFarmArea_GAP_isTipPruning,RehabFarmArea_GAP_isShapePruning,RehabFarmArea_GAP_isAccessPruning,RehabFarmArea_GAP_isSanitaryPruning,RehabFarmArea_GAP_isMaintenancePruning,RehabFarmArea_GAP_isUsingApprTools,RehabFarmArea_GAP_isFarmRecordKeeping,RehabFarmArea_GAP_isChupon,RehabFarmArea_GAP_isFoliarFertilizer,RehabFarmArea_GAP_isSealing,RehabFarmArea_GAP_isObservance,RehabFarmArea_GAP_isGrafting,RehabFarmArea_GAP_ValidatorName,RehabFarmArea_GAP_DateValidation,RehabFarmArea_GAP_ValidationRemarks,AccessToProdInvestment_TypeOfProdInvestments,AccessToProdInvestment_ValidatorName,AccessToProdInvestment_DateValidation,AccessToProdInvestment_ValidationRemarks,filename))
                
                insert=db.do(querycsv)
            
            if(insert["response"]=="error"):
                flash(f"An error occured!", "error")
                print(str(insert))
                print(sheet.name)
            else:
                flash(f"The file was imported successfully!", "success")
            print("DONE HERE!!!!!!!!!!!!")
            return "done"
        
        elif impType == "fo":
            if "DDN_FOs" in wb.sheetnames:
                sheet = wb["DDN_FOs"]
            else:
                flash("Invalid file template!", "error")
                return "done:S"

            data = [list(row) for row in sheet.iter_rows(values_only=True)]
            
            for row in data[9:]:
                if not any(cell and str(cell).strip() for cell in row):
                    continue
                
                UploadedBy = session["USER_DATA"][0]['id']
                ImpUnits_RCU = row[2]
                ImpUnits_PCU = row[3]
                DIP_Name = row[4]
                DIP_Commodity = row[5]
                FOInfo_FOName = row[6]
                FOInfo_Address_Region = row[7]
                FOInfo_Address_Province = row[8]
                FOInfo_Address_Municipality = row[9]
                FOInfo_Address_Brgy = row[10]
                FOInfo_Address_Street = row[11]
                FOInfo_Head_Fname = row[12]
                FOInfo_Head_Mname = row[13]
                FOInfo_Head_Lname = row[14]
                FOInfo_Head_Extname = row[15]
                FOInfo_Head_Sex = row[16]
                FOInfo_Head_Sector = row[17]
                AcknowOfFarmExp_TargetArea = row[18]
                AcknowOfFarmExp_ItemName = row[19]
                AcknowOfFarmExp_QtyReceived = row[20]
                AcknowOfFarmExp_DateReceived = row[21] if row[21] else None
                AcknowOfFarmExp_FarmerRep_Fname = row[22]
                AcknowOfFarmExp_FarmerRep_Mname = row[23]
                AcknowOfFarmExp_FarmerRep_Lname = row[24]
                AcknowOfFarmExp_FarmerRep_Extname = row[25]
                AcknowOfFarmExp_FarmerRep_Designation = row[26]
                AcknowOfFarmExp_Witness1_ValidatorName = row[27]
                AcknowOfFarmExp_Witness1_ValidationDate = row[28] if row[28] else None 
                AcknowOfFarmExp_Witness1_Designation = row[29]
                AcknowOfFarmExp_Witness1_DateWitnessed = row[30] if row[30] else None
                AcknowOfFarmExp_Witness2_ValidatorName = row[31]
                AcknowOfFarmExp_Witness2_ValidationDate = row[32]
                AcknowOfFarmExp_Witness2_Remarks = row[33]
                AcknowOfFarmExp_Witness1Raw_ValidatorName = row[34]
                AcknowOfFarmExp_Witness1Raw_ValidationDate = row[35] if row[35] else None
                AcknowOfFarmExp_Witness1Raw_Designation = row[36]
                AcknowOfFarmExp_Witness1Raw_DateWitnessed = row[37] if row[37] else None
                AcknowOfFarmExp_Witness2Raw_ValidatorName = row[38]
                AcknowOfFarmExp_Witness2Raw_ValidationDate = row[39] if row[39] else None
                AcknowOfFarmExp_Witness2Raw_Remarks = row[40]
                AcknowOfFarmInten_TargetArea = row[41]
                AcknowOfFarmInten_ItemName = row[42]
                AcknowOfFarmInten_QtyReceived = row[43]
                AcknowOfFarmInten_DateReceived = row[44] if row[44] else None
                AcknowOfFarmInten_FarmerRep_Fname = row[45]
                AcknowOfFarmInten_FarmerRep_Mname = row[46]
                AcknowOfFarmInten_FarmerRep_Lname = row[47]
                AcknowOfFarmInten_FarmerRep_Extname = row[48]
                AcknowOfFarmInten_FarmerRep_Designation = row[49]
                AcknowOfFarmInten_Withness1_ValidatorName = row[50]
                AcknowOfFarmInten_Withness1_ValidationDate = row[51]
                AcknowOfFarmInten_Withness1_Designation = row[52]
                AcknowOfFarmInten_Withness1_DateWitnessed = row[53] if row[53] else None
                AcknowOfFarmInten_Withness2_ValidatorName = row[54]
                AcknowOfFarmInten_Withness2_ValidationDate = row[55] if row[55] else None
                AcknowOfFarmInten_Withness2_Remarks = row[56]
                AcknowOfFarmRehab_TargetArea = row[57]
                AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw = row[58]
                AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw = row[59]
                AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer = row[60]
                AcknowOfFarmRehab_DateReceived = row[61] if row[61] else None
                AcknowOfFarmRehab_FarmerRep_Fname = row[62]
                AcknowOfFarmRehab_FarmerRep_Mname = row[63]
                AcknowOfFarmRehab_FarmerRep_Lname = row[64]
                AcknowOfFarmRehab_FarmerRep_Extname = row[65]
                AcknowOfFarmRehab_FarmerRep_Designation = row[66]
                AcknowOfFarmRehab_ValidatorName = row[67]
                AcknowOfFarmRehab_ValidationDate = row[68]
                AcknowOfFarmRehab_Remarks = row[69]
                AcknowOfProd_TypeOfProdInv = row[70]
                AcknowOfProd_QtyItemReceived = row[71]
                AcknowOfProd_DateReceived = row[72] if row[72] else None
                AcknowOfProd_FarmerRep_Fname = row[73]
                AcknowOfProd_FarmerRep_Mname = row[74]
                AcknowOfProd_FarmerRep_Lname = row[75]
                AcknowOfProd_FarmerRep_Extname = row[76]
                AcknowOfProd_FarmerRep_Designation = row[77]
                AcknowOfProd_ValidatorName = row[78]
                AcknowOfProd_ValidationDate = row[79] if row[79] else None
                AcknowOfProd_Remarks = row[80]
                filename = UPLOAD_NAME

                querycsv = ("INSERT INTO mg_implementation_fo (UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,FOInfo_FOName,FOInfo_Address_Region,FOInfo_Address_Province,FOInfo_Address_Municipality,FOInfo_Address_Brgy,FOInfo_Address_Street,FOInfo_Head_Fname,FOInfo_Head_Mname,FOInfo_Head_Lname,FOInfo_Head_Extname,FOInfo_Head_Sex,FOInfo_Head_Sector,AcknowOfFarmExp_TargetArea,AcknowOfFarmExp_ItemName,AcknowOfFarmExp_QtyReceived,AcknowOfFarmExp_DateReceived,AcknowOfFarmExp_FarmerRep_Fname,AcknowOfFarmExp_FarmerRep_Mname,AcknowOfFarmExp_FarmerRep_Lname,AcknowOfFarmExp_FarmerRep_Extname,AcknowOfFarmExp_FarmerRep_Designation,AcknowOfFarmExp_Witness1_ValidatorName,AcknowOfFarmExp_Witness1_ValidationDate,AcknowOfFarmExp_Witness1_Designation,AcknowOfFarmExp_Witness1_DateWitnessed,AcknowOfFarmExp_Witness2_ValidatorName,AcknowOfFarmExp_Witness2_ValidationDate,AcknowOfFarmExp_Witness2_Remarks,AcknowOfFarmExp_Witness1Raw_ValidatorName,AcknowOfFarmExp_Witness1Raw_ValidationDate,AcknowOfFarmExp_Witness1Raw_Designation,AcknowOfFarmExp_Witness1Raw_DateWitnessed,AcknowOfFarmExp_Witness2Raw_ValidatorName,AcknowOfFarmExp_Witness2Raw_ValidationDate,AcknowOfFarmExp_Witness2Raw_Remarks,AcknowOfFarmInten_TargetArea,AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived,AcknowOfFarmInten_DateReceived,AcknowOfFarmInten_FarmerRep_Fname,AcknowOfFarmInten_FarmerRep_Mname,AcknowOfFarmInten_FarmerRep_Lname,AcknowOfFarmInten_FarmerRep_Extname,AcknowOfFarmInten_FarmerRep_Designation,AcknowOfFarmInten_Withness1_ValidatorName,AcknowOfFarmInten_Withness1_ValidationDate,AcknowOfFarmInten_Withness1_Designation,AcknowOfFarmInten_Withness1_DateWitnessed,AcknowOfFarmInten_Withness2_ValidatorName,AcknowOfFarmInten_Withness2_ValidationDate,AcknowOfFarmInten_Withness2_Remarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_FarmerRep_Fname,AcknowOfFarmRehab_FarmerRep_Mname,AcknowOfFarmRehab_FarmerRep_Lname,AcknowOfFarmRehab_FarmerRep_Extname,AcknowOfFarmRehab_FarmerRep_Designation,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_ValidationDate,AcknowOfFarmRehab_Remarks,AcknowOfProd_TypeOfProdInv,AcknowOfProd_QtyItemReceived,AcknowOfProd_DateReceived,AcknowOfProd_FarmerRep_Fname,AcknowOfProd_FarmerRep_Mname,AcknowOfProd_FarmerRep_Lname,AcknowOfProd_FarmerRep_Extname,AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,filename) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".
                format(UploadedBy,ImpUnits_RCU,ImpUnits_PCU,DIP_Name,DIP_Commodity,FOInfo_FOName,FOInfo_Address_Region,FOInfo_Address_Province,FOInfo_Address_Municipality,FOInfo_Address_Brgy,FOInfo_Address_Street,FOInfo_Head_Fname,FOInfo_Head_Mname,FOInfo_Head_Lname,FOInfo_Head_Extname,FOInfo_Head_Sex,FOInfo_Head_Sector,AcknowOfFarmExp_TargetArea,AcknowOfFarmExp_ItemName,AcknowOfFarmExp_QtyReceived,AcknowOfFarmExp_DateReceived,AcknowOfFarmExp_FarmerRep_Fname,AcknowOfFarmExp_FarmerRep_Mname,AcknowOfFarmExp_FarmerRep_Lname,AcknowOfFarmExp_FarmerRep_Extname,AcknowOfFarmExp_FarmerRep_Designation,AcknowOfFarmExp_Witness1_ValidatorName,AcknowOfFarmExp_Witness1_ValidationDate,AcknowOfFarmExp_Witness1_Designation,AcknowOfFarmExp_Witness1_DateWitnessed,AcknowOfFarmExp_Witness2_ValidatorName,AcknowOfFarmExp_Witness2_ValidationDate,AcknowOfFarmExp_Witness2_Remarks,AcknowOfFarmExp_Witness1Raw_ValidatorName,AcknowOfFarmExp_Witness1Raw_ValidationDate,AcknowOfFarmExp_Witness1Raw_Designation,AcknowOfFarmExp_Witness1Raw_DateWitnessed,AcknowOfFarmExp_Witness2Raw_ValidatorName,AcknowOfFarmExp_Witness2Raw_ValidationDate,AcknowOfFarmExp_Witness2Raw_Remarks,AcknowOfFarmInten_TargetArea,AcknowOfFarmInten_ItemName,AcknowOfFarmInten_QtyReceived,AcknowOfFarmInten_DateReceived,AcknowOfFarmInten_FarmerRep_Fname,AcknowOfFarmInten_FarmerRep_Mname,AcknowOfFarmInten_FarmerRep_Lname,AcknowOfFarmInten_FarmerRep_Extname,AcknowOfFarmInten_FarmerRep_Designation,AcknowOfFarmInten_Withness1_ValidatorName,AcknowOfFarmInten_Withness1_ValidationDate,AcknowOfFarmInten_Withness1_Designation,AcknowOfFarmInten_Withness1_DateWitnessed,AcknowOfFarmInten_Withness2_ValidatorName,AcknowOfFarmInten_Withness2_ValidationDate,AcknowOfFarmInten_Withness2_Remarks,AcknowOfFarmRehab_TargetArea,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_HighPrunerSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_MiniChainSaw,AcknowOfFarmRehab_QtyOfFarmRehabToolsRec_KnapsackSprayer,AcknowOfFarmRehab_DateReceived,AcknowOfFarmRehab_FarmerRep_Fname,AcknowOfFarmRehab_FarmerRep_Mname,AcknowOfFarmRehab_FarmerRep_Lname,AcknowOfFarmRehab_FarmerRep_Extname,AcknowOfFarmRehab_FarmerRep_Designation,AcknowOfFarmRehab_ValidatorName,AcknowOfFarmRehab_ValidationDate,AcknowOfFarmRehab_Remarks,AcknowOfProd_TypeOfProdInv,AcknowOfProd_QtyItemReceived,AcknowOfProd_DateReceived,AcknowOfProd_FarmerRep_Fname,AcknowOfProd_FarmerRep_Mname,AcknowOfProd_FarmerRep_Lname,AcknowOfProd_FarmerRep_Extname,AcknowOfProd_FarmerRep_Designation,AcknowOfProd_ValidatorName,AcknowOfProd_ValidationDate,AcknowOfProd_Remarks,filename))
                
                insert=db.do(querycsv)
        
            if(insert["response"]=="error"):
                flash(f"An error occured!", "error")
                print(str(insert))
                print(sheet.name)
            else:
                flash(f"The file was imported successfully!", "success")
            return "done"