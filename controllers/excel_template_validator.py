import os
from zipfile import BadZipFile

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


DOWNLOADABLES_DIR = os.path.abspath(
	os.path.join(os.path.dirname(__file__), "..", "assets", "downloadables")
)


def normalize_header_value(value):
	if value is None:
		return ""
	if isinstance(value, float) and value.is_integer():
		value = int(value)
	return " ".join(str(value).split()).casefold()


def _xlrd_sheet(book, expected_name):
	expected_name = expected_name.strip().casefold()
	for sheet_name in book.sheet_names():
		if sheet_name.strip().casefold() == expected_name:
			return book.sheet_by_name(sheet_name)
	return None


def _xlrd_signature(sheet, header_row_count):
	return [
		[
			normalize_header_value(sheet.cell_value(row_index, column_index))
			for column_index in range(sheet.ncols)
		]
		for row_index in range(header_row_count)
	]


def _open_xlrd_upload(upload_path):
	if hasattr(upload_path, "read"):
		original_position = upload_path.tell()
		upload_contents = upload_path.read()
		upload_path.seek(original_position)
		return xlrd.open_workbook(file_contents=upload_contents)
	return xlrd.open_workbook(upload_path)


def validate_xlrd_template_sheets(upload_path, template_filenames, sheet_specs):
	if isinstance(template_filenames, str):
		template_filenames = [template_filenames]

	try:
		upload_book = _open_xlrd_upload(upload_path)
	except (OSError, TypeError, ValueError, xlrd.XLRDError):
		return False, "Invalid Excel file. Please upload the downloaded system template."

	for template_filename in template_filenames:
		try:
			template_book = xlrd.open_workbook(
				os.path.join(DOWNLOADABLES_DIR, template_filename)
			)
		except (OSError, ValueError, xlrd.XLRDError):
			continue

		template_matches = True
		for sheet_name, header_row_count in sheet_specs:
			upload_sheet = _xlrd_sheet(upload_book, sheet_name)
			template_sheet = _xlrd_sheet(template_book, sheet_name)

			if (
				upload_sheet is None
				or template_sheet is None
				or upload_sheet.nrows < header_row_count
				or upload_sheet.ncols != template_sheet.ncols
				or _xlrd_signature(upload_sheet, header_row_count)
				!= _xlrd_signature(template_sheet, header_row_count)
			):
				template_matches = False
				break

		if template_matches:
			return True, ""

	return (
		False,
		"Excel columns do not match the system template. "
		"One or more column names or their order are incorrect. "
		"Please download and use the latest system template.",
	)


def validate_xlrd_template(upload_path, template_filenames, sheet_name, header_row_count):
	return validate_xlrd_template_sheets(
		upload_path,
		template_filenames,
		[(sheet_name, header_row_count)],
	)


def _openpyxl_signature(sheet, header_row_count, column_count):
	return [
		[
			normalize_header_value(sheet.cell(row=row_index, column=column_index).value)
			for column_index in range(1, column_count + 1)
		]
		for row_index in range(1, header_row_count + 1)
	]


def validate_openpyxl_template(
	upload_path,
	template_filename,
	sheet_name,
	header_row_count,
):
	try:
		upload_book = load_workbook(upload_path, data_only=False, read_only=True)
		template_book = load_workbook(
			os.path.join(DOWNLOADABLES_DIR, template_filename),
			data_only=False,
			read_only=True,
		)
	except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException):
		return False, "Invalid Excel file. Please upload the downloaded system template."

	if sheet_name not in upload_book.sheetnames or sheet_name not in template_book.sheetnames:
		upload_book.close()
		template_book.close()
		return False, "Excel format does not match the system template."

	upload_sheet = upload_book[sheet_name]
	template_sheet = template_book[sheet_name]
	column_count = template_sheet.max_column

	if (
		upload_sheet.max_column != column_count
		or _openpyxl_signature(upload_sheet, header_row_count, column_count)
		!= _openpyxl_signature(template_sheet, header_row_count, column_count)
	):
		upload_book.close()
		template_book.close()
		return (
			False,
			"Excel columns do not match the system template. "
			"One or more column names or their order are incorrect. "
			"Please download and use the latest system template.",
		)

	upload_book.close()
	template_book.close()
	return True, ""
